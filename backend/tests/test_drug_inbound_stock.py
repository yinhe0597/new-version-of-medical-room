import io
import unittest
from datetime import date, datetime, time, timedelta

from backend.app import create_app, db
from backend.app.models import (
    DailyStockSnapshot,
    Drug,
    DrugStockGroup,
    InventoryRecord,
    Patient,
    Payment,
    PrescriptionItem,
    User,
    Visit,
    INVENTORY_OPERATION_ADJUSTMENT,
    INVENTORY_OPERATION_DISPENSE,
    INVENTORY_OPERATION_INBOUND,
    INVENTORY_OPERATION_REVERSAL,
    VISIT_STATUS_COMPLETED,
    utcnow,
)
from backend.app.api.nurse import _compute_monthly_report
from backend.app.services.drug_stock import recompute_variant_stocks
from openpyxl import load_workbook


class TestConfig:
    TESTING = True
    SECRET_KEY = "test"
    JWT_SECRET_KEY = "test-jwt-secret-key-at-least-32-bytes"
    SQLALCHEMY_DATABASE_URI = "sqlite:///:memory:"
    SQLALCHEMY_TRACK_MODIFICATIONS = False


class DrugInboundStockTestCase(unittest.TestCase):
    def setUp(self):
        self.app = create_app(TestConfig)
        self.ctx = self.app.app_context()
        self.ctx.push()
        db.create_all()
        self.client = self.app.test_client()

        self.nurse = User(username="nurse", real_name="护士", role="nurse")
        self.nurse.set_password("123456")
        self.doctor = User(username="doctor", real_name="医生", role="doctor")
        self.doctor.set_password("123456")
        db.session.add_all([self.nurse, self.doctor])
        db.session.commit()

        resp = self.client.post("/api/auth/login", json={"username": "nurse", "password": "123456"})
        self.assertEqual(resp.status_code, 200)
        token = resp.get_json()["access_token"]
        self.headers = {"Authorization": f"Bearer {token}"}

    def _create_group(self, name, batch_no, inbound_quantity=2):
        response = self.client.post(
            "/api/nurse/inbound",
            json={
                "type": 1,
                "name": name,
                "batch_no": batch_no,
                "pack_specification": "20 mg×100粒/瓶",
                "pack_price": 10.0,
                "purchase_price": 6.0,
                "inbound_quantity": inbound_quantity,
                "retail_enabled": True,
                "min_sale_unit": "2粒",
                "min_sale_price": 0.30,
            },
            headers=self.headers,
        )
        self.assertEqual(response.status_code, 201)
        data = response.get_json()["data"]
        group = DrugStockGroup.query.filter_by(group_code=data["group_code"]).one()
        return group, group.pack_drug, group.retail_drug

    def _add_completed_item(self, drug, quantity, report_day):
        patient = Patient(name="月报患者", gender="男", student_id=f"S{drug.id}")
        db.session.add(patient)
        db.session.flush()
        paid_at = datetime.combine(report_day, time(hour=12)) - timedelta(hours=8)
        visit = Visit(
            patient_id=patient.id,
            doctor_id=self.doctor.id,
            timestamp=paid_at,
            status=VISIT_STATUS_COMPLETED,
        )
        db.session.add(visit)
        db.session.flush()
        db.session.add_all([
            PrescriptionItem(
                visit_id=visit.id,
                drug_id=drug.id,
                quantity=quantity,
                price_at_visit=drug.price,
                amount=float(drug.price or 0) * quantity,
                is_scattered=False,
            ),
            Payment(
                visit_id=visit.id,
                nurse_id=self.nurse.id,
                amount=float(drug.price or 0) * quantity,
                payment_date=paid_at,
                payment_method="cash",
            ),
        ])
        db.session.commit()

    def tearDown(self):
        db.session.remove()
        db.drop_all()
        self.ctx.pop()

    def test_inbound_creates_pack_and_retail_and_group(self):
        payload = {
            "type": 1,
            "name": "药品A",
            "batch_no": "B001",
            "pack_specification": "20 mg×100粒/瓶",
            "pack_price": 10.0,
            "purchase_price": 6.0,
            "inbound_quantity": 3,
            "retail_enabled": True,
            "min_sale_unit": "2粒",
            "min_sale_price": 0.30,
        }
        resp = self.client.post("/api/nurse/inbound", json=payload, headers=self.headers)
        self.assertEqual(resp.status_code, 201)
        data = resp.get_json()["data"]
        self.assertIn("group_code", data)

        drugs = Drug.query.filter(Drug.batch_no == "B001").all()
        self.assertEqual(len(drugs), 2)

        group = DrugStockGroup.query.filter_by(group_code=data["group_code"]).first()
        self.assertIsNotNone(group)
        self.assertEqual(group.total_units, 300)
        self.assertEqual(group.pack_amount, 100)
        self.assertEqual(group.retail_amount, 2)

        pack = db.session.get(Drug, data["pack_drug_id"])
        retail = db.session.get(Drug, data["retail_drug_id"])
        self.assertEqual(pack.stock, 3)
        self.assertEqual(retail.stock, 150)
        self.assertEqual(pack.purchase_price, 6.0)
        self.assertAlmostEqual(retail.purchase_price, 0.12, places=6)
        pack_record = InventoryRecord.query.filter_by(drug_id=pack.id).one()
        retail_record = InventoryRecord.query.filter_by(drug_id=retail.id).one()
        self.assertEqual((pack_record.old_stock, pack_record.new_stock), (0, 3))
        self.assertEqual((retail_record.old_stock, retail_record.new_stock), (0, 150))
        self.assertEqual(pack_record.operation_type, INVENTORY_OPERATION_INBOUND)
        self.assertEqual(retail_record.operation_type, INVENTORY_OPERATION_INBOUND)

        today = date.today().isoformat()
        report, error = _compute_monthly_report(today, today)
        self.assertIsNone(error)
        grouped_rows = [row for row in report if row["drug_id"] in {pack.id, retail.id}]
        self.assertEqual(len(grouped_rows), 1)
        self.assertEqual(grouped_rows[0]["unit"], "粒")
        self.assertEqual(grouped_rows[0]["inbound"], 300)
        self.assertEqual(grouped_rows[0]["closing_stock"], 300)
        self.assertAlmostEqual(grouped_rows[0]["inbound_amount"], 18.0, places=6)

        listing = self.client.get("/api/nurse/drugs?size=20", headers=self.headers)
        self.assertEqual(listing.status_code, 200)
        listed_group_rows = [
            row for row in listing.get_json()["data"] if row["stock_group_code"] == group.group_code
        ]
        self.assertEqual(len(listed_group_rows), 2)
        self.assertTrue(all(row["group_total_units"] == 300 for row in listed_group_rows))
        self.assertTrue(all(row["group_pack_amount"] == 100 for row in listed_group_rows))

    def test_group_inventory_uses_group_conversion_for_either_variant(self):
        group, _pack, _retail = self._create_group("联合盘点药", "GROUP-ADJUST", 3)

        invalid = self.client.post(
            "/api/nurse/inventory/group",
            json={
                "group_code": group.group_code,
                "actual_packs": 1,
                "actual_retail_units": 50,
                "remark": "测试折算上限",
            },
            headers=self.headers,
        )
        self.assertEqual(invalid.status_code, 400)
        self.assertEqual(group.total_units, 300)

        oversized = self.client.post(
            "/api/nurse/inventory/group",
            json={
                "group_code": group.group_code,
                "actual_packs": 30_000_000,
                "actual_retail_units": 0,
                "remark": "超大库存",
            },
            headers=self.headers,
        )
        self.assertEqual(oversized.status_code, 400)
        self.assertEqual(group.total_units, 300)

        response = self.client.post(
            "/api/nurse/inventory/group",
            json={
                "group_code": group.group_code,
                "actual_packs": 1,
                "actual_retail_units": 1,
                "remark": "实物盘点",
            },
            headers=self.headers,
        )
        self.assertEqual(response.status_code, 200)
        db.session.refresh(group)
        self.assertEqual(group.total_units, 102)
        self.assertEqual(group.pack_drug.stock, 1)
        self.assertEqual(group.retail_drug.stock, 51)
        adjustment = InventoryRecord.query.filter_by(
            drug_id=group.retail_drug_id,
            operation_type=INVENTORY_OPERATION_ADJUSTMENT,
        ).one()
        self.assertEqual((adjustment.old_stock, adjustment.new_stock), (150, 51))

    def test_group_pack_outbound_report_balances_in_base_units(self):
        group, pack, retail = self._create_group("整装出库药", "PACK-OUT", 2)
        report_day = date.today() - timedelta(days=2)
        group.total_units = 100
        stocks = recompute_variant_stocks(100, group.pack_amount, group.retail_amount)
        pack.stock = stocks["pack_stock"]
        retail.stock = stocks["retail_stock"]
        report_boundary = datetime.combine(report_day, time()) - timedelta(hours=8)
        db.session.add_all([
            DailyStockSnapshot(
                drug_id=retail.id, date=report_day, stock=100,
                created_at=report_boundary,
            ),
            DailyStockSnapshot(
                drug_id=retail.id, date=report_day + timedelta(days=1), stock=50,
                created_at=report_boundary + timedelta(days=1),
            ),
        ])
        db.session.commit()
        self._add_completed_item(pack, 1, report_day)

        report, error = _compute_monthly_report(report_day.isoformat(), report_day.isoformat())
        self.assertIsNone(error)
        row = next(item for item in report if item.get("stock_group_code") == group.group_code)
        self.assertEqual((row["opening_stock"], row["outbound"], row["closing_stock"]), (200, 100, 100))
        self.assertEqual(
            row["opening_stock"] + row["inbound"] + row["adjustment"] - row["outbound"],
            row["closing_stock"],
        )

    def test_group_retail_outbound_crossing_pack_boundary_balances(self):
        group, pack, retail = self._create_group("零售跨界药", "RETAIL-OUT", 2)
        report_day = date.today() - timedelta(days=2)
        group.total_units = 98
        stocks = recompute_variant_stocks(98, group.pack_amount, group.retail_amount)
        pack.stock = stocks["pack_stock"]
        retail.stock = stocks["retail_stock"]
        report_boundary = datetime.combine(report_day, time()) - timedelta(hours=8)
        db.session.add_all([
            DailyStockSnapshot(
                drug_id=retail.id, date=report_day, stock=51,
                created_at=report_boundary,
            ),
            DailyStockSnapshot(
                drug_id=retail.id, date=report_day + timedelta(days=1), stock=49,
                created_at=report_boundary + timedelta(days=1),
            ),
        ])
        db.session.commit()
        self._add_completed_item(retail, 2, report_day)

        report, error = _compute_monthly_report(report_day.isoformat(), report_day.isoformat())
        self.assertIsNone(error)
        row = next(item for item in report if item.get("stock_group_code") == group.group_code)
        self.assertEqual((row["opening_stock"], row["outbound"], row["closing_stock"]), (102, 4, 98))
        self.assertEqual(row["opening_stock"] - row["outbound"], row["closing_stock"])

    def test_historical_report_uses_next_day_snapshot_and_keeps_inactive_drug(self):
        report_day = date.today() - timedelta(days=2)
        drug = Drug(
            name="已停用历史药",
            type=1,
            specification="1盒",
            unit="盒",
            price=10,
            purchase_price=5,
            stock=9,
            status=0,
        )
        db.session.add(drug)
        db.session.flush()
        db.session.add_all([
            DailyStockSnapshot(
                drug_id=drug.id, date=report_day, stock=10,
                created_at=datetime.combine(report_day, time()) - timedelta(hours=8),
            ),
            DailyStockSnapshot(
                drug_id=drug.id, date=report_day + timedelta(days=1), stock=9,
                created_at=datetime.combine(report_day + timedelta(days=1), time()) - timedelta(hours=8),
            ),
        ])
        db.session.commit()
        self._add_completed_item(drug, 1, report_day)

        report, error = _compute_monthly_report(report_day.isoformat(), report_day.isoformat())
        self.assertIsNone(error)
        row = next(item for item in report if item["drug_id"] == drug.id)
        self.assertEqual((row["opening_stock"], row["outbound"], row["closing_stock"]), (10, 1, 9))

    def test_cross_day_reversal_keeps_both_days_balanced(self):
        first_day = date.today() - timedelta(days=3)
        second_day = first_day + timedelta(days=1)
        boundary = lambda value: datetime.combine(value, time()) - timedelta(hours=8)
        drug = Drug(
            name="跨日撤销药", type=1, specification="1盒", unit="盒",
            price=10, purchase_price=5, stock=10, status=1,
        )
        db.session.add(drug)
        db.session.flush()
        visit = Visit(
            patient_id=None,
            doctor_id=self.doctor.id,
            status="revoked",
            revoked_by=self.nurse.id,
            revoked_at=boundary(second_day) + timedelta(hours=4),
        )
        db.session.add(visit)
        db.session.flush()
        db.session.add_all([
            DailyStockSnapshot(
                drug_id=drug.id, date=first_day, stock=10,
                created_at=boundary(first_day),
            ),
            DailyStockSnapshot(
                drug_id=drug.id, date=second_day, stock=9,
                created_at=boundary(second_day),
            ),
            DailyStockSnapshot(
                drug_id=drug.id, date=second_day + timedelta(days=1), stock=10,
                created_at=boundary(second_day + timedelta(days=1)),
            ),
            InventoryRecord(
                drug_id=drug.id,
                nurse_id=self.nurse.id,
                visit_id=visit.id,
                old_stock=10,
                new_stock=9,
                operation_type=INVENTORY_OPERATION_DISPENSE,
                timestamp=boundary(first_day) + timedelta(hours=4),
            ),
            InventoryRecord(
                drug_id=drug.id,
                nurse_id=self.nurse.id,
                visit_id=visit.id,
                old_stock=9,
                new_stock=10,
                operation_type=INVENTORY_OPERATION_REVERSAL,
                timestamp=boundary(second_day) + timedelta(hours=4),
            ),
        ])
        db.session.commit()

        first_report, error = _compute_monthly_report(first_day.isoformat(), first_day.isoformat())
        self.assertIsNone(error)
        first_row = next(item for item in first_report if item["drug_id"] == drug.id)
        self.assertEqual(
            (first_row["opening_stock"], first_row["outbound"], first_row["closing_stock"]),
            (10, 1, 9),
        )

        second_report, error = _compute_monthly_report(second_day.isoformat(), second_day.isoformat())
        self.assertIsNone(error)
        second_row = next(item for item in second_report if item["drug_id"] == drug.id)
        self.assertEqual(
            (second_row["opening_stock"], second_row["adjustment"], second_row["closing_stock"]),
            (9, 1, 10),
        )

    def test_delayed_snapshot_is_normalized_to_midnight_cutoff(self):
        report_day = date.today() - timedelta(days=3)
        next_day = report_day + timedelta(days=1)
        report_boundary = datetime.combine(report_day, time()) - timedelta(hours=8)
        next_boundary = datetime.combine(next_day, time()) - timedelta(hours=8)
        drug = Drug(
            name="零点延迟快照药", type=1, specification="1盒", unit="盒",
            price=10, purchase_price=5, stock=9, status=1,
        )
        db.session.add(drug)
        db.session.flush()
        db.session.add_all([
            DailyStockSnapshot(
                drug_id=drug.id,
                date=report_day,
                stock=10,
                created_at=report_boundary,
            ),
            DailyStockSnapshot(
                drug_id=drug.id,
                date=next_day,
                stock=9,
                created_at=next_boundary + timedelta(milliseconds=500),
            ),
            InventoryRecord(
                drug_id=drug.id,
                nurse_id=self.nurse.id,
                old_stock=10,
                new_stock=9,
                operation_type=INVENTORY_OPERATION_DISPENSE,
                timestamp=next_boundary + timedelta(milliseconds=200),
            ),
        ])
        db.session.commit()

        report, error = _compute_monthly_report(report_day.isoformat(), report_day.isoformat())
        self.assertIsNone(error)
        row = next(item for item in report if item["drug_id"] == drug.id)
        self.assertEqual((row["opening_stock"], row["outbound"], row["closing_stock"]), (10, 0, 10))

    def test_structured_operation_type_overrides_free_text_remark(self):
        now = utcnow()
        inbound_drug = Drug(
            name="结构化入库", type=1, specification="1盒", unit="盒",
            price=10, purchase_price=5, stock=5, status=1,
        )
        adjusted_drug = Drug(
            name="结构化盘点", type=1, specification="1盒", unit="盒",
            price=10, purchase_price=5, stock=7, status=1,
        )
        db.session.add_all([inbound_drug, adjusted_drug])
        db.session.flush()
        db.session.add_all([
            InventoryRecord(
                drug_id=inbound_drug.id,
                nurse_id=self.nurse.id,
                old_stock=0,
                new_stock=5,
                operation_type=INVENTORY_OPERATION_INBOUND,
                remark="不是入库前缀",
                timestamp=now,
            ),
            InventoryRecord(
                drug_id=adjusted_drug.id,
                nurse_id=self.nurse.id,
                old_stock=0,
                new_stock=7,
                operation_type=INVENTORY_OPERATION_ADJUSTMENT,
                remark="入库字样不能改变结构化类型",
                timestamp=now,
            ),
        ])
        db.session.commit()

        today = date.today().isoformat()
        report, error = _compute_monthly_report(today, today)
        self.assertIsNone(error)
        inbound_row = next(item for item in report if item["drug_id"] == inbound_drug.id)
        adjusted_row = next(item for item in report if item["drug_id"] == adjusted_drug.id)
        self.assertEqual((inbound_row["inbound"], inbound_row["adjustment"]), (5, 0))
        self.assertEqual((adjusted_row["inbound"], adjusted_row["adjustment"]), (0, 7))

    def test_monthly_export_escapes_spreadsheet_formulas(self):
        drug = Drug(
            name='=HYPERLINK("https://example.invalid")',
            type=1,
            specification="+1盒",
            unit="@盒",
            price=10,
            purchase_price=5,
            stock=1,
            status=1,
        )
        db.session.add(drug)
        db.session.commit()
        today = date.today().isoformat()

        response = self.client.get(
            "/api/nurse/inventory/monthly-report/export",
            query_string={"start_date": today, "end_date": today},
            headers=self.headers,
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data), data_only=False)
        row = next(
            values for values in workbook.active.iter_rows(min_row=2, values_only=True)
            if values[1] and "HYPERLINK" in values[1]
        )
        self.assertTrue(row[1].startswith("'="))
        self.assertTrue(row[3].startswith("'+"))
        self.assertTrue(row[5].startswith("'@"))

    def test_inbound_rejects_low_min_sale_price(self):
        payload = {
            "type": 1,
            "name": "药品B",
            "batch_no": "B002",
            "pack_specification": "20 mg×100粒/瓶",
            "pack_price": 10.0,
            "purchase_price": 6.0,
            "inbound_quantity": 1,
            "retail_enabled": True,
            "min_sale_unit": "1粒",
            "min_sale_price": 0.05,
        }
        resp = self.client.post("/api/nurse/inbound", json=payload, headers=self.headers)
        self.assertEqual(resp.status_code, 400)

    def test_inbound_duplicate_batch_returns_409(self):
        payload = {
            "type": 1,
            "name": "药品C",
            "batch_no": "B003",
            "pack_specification": "20 mg×100粒/瓶",
            "pack_price": 10.0,
            "purchase_price": 6.0,
            "inbound_quantity": 1,
            "retail_enabled": False,
        }
        r1 = self.client.post("/api/nurse/inbound", json=payload, headers=self.headers)
        self.assertEqual(r1.status_code, 201)
        r2 = self.client.post("/api/nurse/inbound", json=payload, headers=self.headers)
        self.assertEqual(r2.status_code, 409)


if __name__ == "__main__":
    unittest.main()

