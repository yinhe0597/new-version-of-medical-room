from flask import request, jsonify, send_file, make_response, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from backend.app import db
from backend.app.api import bp
from backend.app.models import (
    User, Drug, DrugStockGroup, Payment, Visit, PrescriptionItem, Patient, InventoryRecord,
    DailyStockSnapshot, OperationLog,
    INVENTORY_OPERATION_ADJUSTMENT, INVENTORY_OPERATION_INBOUND, INVENTORY_OPERATION_MERGE,
)
from backend.app.utils.decorators import role_required
from backend.app.services.stock_lock import serialized_stock_mutation
from backend.app.services.revenue import allocate_payment_revenue
from backend.app.services.time_utils import (
    format_local_datetime,
    local_naive_to_utc,
    local_now,
    parse_local_datetime,
    utc_naive_to_local,
)
from backend.app.services.bootstrap import validate_password
from backend.app.utils.query import nulls_last_asc
from datetime import datetime, date, timedelta
from sqlalchemy import func
from sqlalchemy.orm import aliased
from sqlalchemy.engine import make_url
import os
import csv
import io
import json
import math
import re
import subprocess
import sqlite3
import tempfile
import zipfile
from sqlalchemy.exc import IntegrityError

from pypinyin import pinyin, Style

MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10MB

PATIENT_TYPES = {
    'student': '学生',
    'staff': '教职工',
    'shop': '商铺员工',
    'temporary': '临时人员',
}

_ID_CARD_RE = re.compile(r"^\d{17}[\dXx]$")

def _is_valid_cn_id_card(id_card: str) -> bool:
    if not isinstance(id_card, str):
        return False
    s = id_card.strip()
    if not s:
        return False
    if not _ID_CARD_RE.match(s):
        return False
    birth = s[6:14]
    try:
        datetime.strptime(birth, "%Y%m%d")
    except Exception:
        return False
    weights = [7, 9, 10, 5, 8, 4, 2, 1, 6, 3, 7, 9, 10, 5, 8, 4, 2]
    mapping = ["1", "0", "X", "9", "8", "7", "6", "5", "4", "3", "2"]
    total = 0
    for i in range(17):
        total += int(s[i]) * weights[i]
    check = mapping[total % 11]
    return s[-1].upper() == check

def _age_from_id_card(id_card: str):
    """从 18 位中国身份证号提取出生日期并计算周岁"""
    if not id_card or len(id_card) != 18:
        return None
    try:
        birth = datetime.strptime(id_card[6:14], "%Y%m%d")
        today = local_now()
        age = today.year - birth.year
        if (today.month, today.day) < (birth.month, birth.day):
            age -= 1
        return age
    except Exception:
        return None

def _check_upload_size(file_storage):
    file_storage.seek(0, 2)
    size = file_storage.tell()
    file_storage.seek(0)
    if size > MAX_UPLOAD_SIZE:
        return False
    return True


def _mask_patient_name(name, finance_view=None):
    """对财务角色脱敏患者姓名"""
    if finance_view is None:
        finance_view = _current_user_is_finance()
    if finance_view and name:
        if len(name) == 1:
            return '*'
        return name[0] + '*' * (len(name) - 1)
    return name or ''


def _current_user_is_finance():
    try:
        user = db.session.get(User, int(get_jwt_identity()))
        return bool(user and user.role == 'finance')
    except (TypeError, ValueError):
        return False

def _name_pinyin_parts(text):
    if not isinstance(text, str) or not text:
        return "", ""
    initials_list = pinyin(text, style=Style.FIRST_LETTER, strict=False)
    initials = "".join([x[0] for x in initials_list if x]).lower()
    full_list = pinyin(text, style=Style.NORMAL, strict=False)
    full = "".join([x[0] for x in full_list if x]).lower()
    return full, initials


def _nonnegative_float(value, field):
    try:
        result = float(value)
    except (TypeError, ValueError):
        raise ValueError(field)
    if not math.isfinite(result) or result < 0:
        raise ValueError(field)
    return result


def _nonnegative_int(value, field):
    if isinstance(value, bool):
        raise ValueError(field)
    try:
        result = int(value)
        numeric = float(value)
    except (TypeError, ValueError, OverflowError):
        raise ValueError(field)
    if not math.isfinite(numeric) or numeric != result or result < 0:
        raise ValueError(field)
    return result


def _mysql_backup_response(uri):
    url = make_url(uri)
    if not url.drivername.startswith("mysql") or not url.database:
        return jsonify({"msg": "当前数据库不是 MySQL"}), 400

    command = [
        "mysqldump",
        "-h", url.host or "127.0.0.1",
        "-P", str(url.port or 3306),
        "-u", url.username or "root",
        "--single-transaction",
        "--default-character-set=utf8mb4",
        url.database,
    ]
    env = os.environ.copy()
    if url.password:
        env["MYSQL_PWD"] = url.password

    fd, backup_path = tempfile.mkstemp(prefix="medical_backup_", suffix=".sql")
    os.close(fd)
    try:
        with open(backup_path, "wb") as output:
            process = subprocess.run(
                command,
                stdout=output,
                stderr=subprocess.PIPE,
                env=env,
                check=False,
                timeout=300,
            )
    except FileNotFoundError:
        os.remove(backup_path)
        return jsonify({"msg": "未找到 mysqldump，请安装 MySQL 客户端工具"}), 500
    except subprocess.TimeoutExpired:
        if os.path.exists(backup_path):
            os.remove(backup_path)
        current_app.logger.error("mysqldump timed out after 300 seconds")
        return jsonify({"msg": "MySQL 备份超时，请检查数据库负载"}), 504
    except OSError:
        if os.path.exists(backup_path):
            os.remove(backup_path)
        current_app.logger.exception("Failed to start mysqldump")
        return jsonify({"msg": "MySQL 备份失败，请检查客户端工具和临时目录"}), 500

    if process.returncode != 0:
        current_app.logger.error(
            "mysqldump failed: %s",
            process.stderr.decode("utf-8", errors="replace"),
        )
        os.remove(backup_path)
        return jsonify({"msg": "MySQL 备份失败，请检查服务和日志"}), 500

    timestamp = local_now().strftime("%Y%m%d_%H%M%S")
    try:
        response = send_file(
            backup_path,
            as_attachment=True,
            download_name=f"{url.database}_backup_{timestamp}.sql",
            mimetype="application/sql",
        )
    except Exception:
        os.remove(backup_path)
        raise
    response.direct_passthrough = False
    response.call_on_close(lambda: os.path.exists(backup_path) and os.remove(backup_path))
    return response


def _create_sqlite_backup(uri):
    url = make_url(uri)
    if not url.drivername.startswith("sqlite") or not url.database or url.database == ":memory:":
        return None, (jsonify({"msg": "当前数据库不是可备份的 SQLite 文件"}), 400)

    db_path = os.path.abspath(url.database)
    if not os.path.isfile(db_path):
        return None, (jsonify({"msg": "数据库文件不存在"}), 500)

    backup_dir = os.path.join(os.path.dirname(db_path), "backups")
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = local_now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = os.path.join(backup_dir, f"backup_{timestamp}.db")
    source = sqlite3.connect(db_path, timeout=15)
    destination = sqlite3.connect(backup_path, timeout=15)
    try:
        source.backup(destination)
    except Exception:
        destination.close()
        source.close()
        if os.path.exists(backup_path):
            os.remove(backup_path)
        raise
    else:
        destination.close()
        source.close()
    return backup_path, None


def _prune_sqlite_backups(backup_path, keep=20):
    backup_dir = os.path.dirname(backup_path)
    backups = sorted(
        (
            os.path.join(backup_dir, name)
            for name in os.listdir(backup_dir)
            if name.startswith("backup_") and name.endswith(".db")
        ),
        key=os.path.getmtime,
        reverse=True,
    )
    for stale_path in backups[keep:]:
        try:
            os.remove(stale_path)
        except OSError:
            current_app.logger.warning("Failed to prune backup: %s", stale_path)

@bp.route('/admin/backup/mysql', methods=['GET'])
@role_required('admin')
def backup_mysql_database():
    return _mysql_backup_response(current_app.config.get('SQLALCHEMY_DATABASE_URI', ''))

@bp.route('/admin/patients/template', methods=['GET'])
@role_required('admin')
def get_patients_template():
    template_type = request.args.get('type', 'student')

    templates = {
        'student': "学号,姓名,性别,手机号码,年级,学院,专业,班级,年龄,辅导员姓名\n2024001,张三,男,,2024级,计算机学院,软件工程,软件一班,19,王老师\n2024002,李四,女,13912345678,2024级,外国语学院,英语,英语二班,20,赵老师\n",
        'staff': "姓名,性别,身份证号,手机号码,所在单位\n张三,男,110101199001011234,13800138000,计算机学院\n李四,女,110101198512121234,13900139000,外国语学院\n",
        'shop': "姓名,性别,身份证号,手机号码,商铺名称\n王五,男,110101198805053456,13700137000,校园超市\n赵六,女,110101199208084567,13600136000,学生食堂\n",
    }

    csv_content = templates.get(template_type, templates['student'])
    filename_map = {
        'student': 'patients_template_student.csv',
        'staff': 'patients_template_staff.csv',
        'shop': 'patients_template_shop.csv',
    }
    response = make_response(csv_content.encode('utf-8-sig'))
    response.headers["Content-Disposition"] = f"attachment; filename={filename_map.get(template_type, 'patients_template.csv')}"
    response.headers["Content-type"] = "text/csv"
    return response

@bp.route('/admin/patients/import', methods=['POST'])
@role_required('admin')
def import_patients():
    if 'file' not in request.files:
        return jsonify({"msg": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"msg": "No selected file"}), 400
    if not file.filename.endswith('.csv'):
        return jsonify({"msg": "Only CSV files are allowed"}), 400
    if not _check_upload_size(file):
        return jsonify({"msg": "File too large, max 10MB"}), 400

    import_type = request.args.get('type', 'student')

    try:
        file_content = file.stream.read()
        try:
            decoded_content = file_content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                decoded_content = file_content.decode("gbk")
            except UnicodeDecodeError:
                decoded_content = file_content.decode("utf-8")

        stream = io.StringIO(decoded_content, newline=None)
        csv_input = csv.DictReader(stream)

        success_count = 0
        error_count = 0

        def pick(row, *keys):
            for k in keys:
                if k in row:
                    return row.get(k)
            return None

        def sanitize(val):
            if val and val[0] in ('=', '+', '-', '@'):
                return "'" + val
            return val

        if import_type == 'staff':
            # 教职工导入: 姓名,性别,身份证号,手机号码,所在单位
            for row in csv_input:
                name = sanitize((pick(row, 'name', '姓名') or '').strip())
                gender = (pick(row, 'gender', '性别') or '').strip()
                id_card = (pick(row, 'id_card', '身份证号') or '').strip()
                phone = (pick(row, 'phone', '手机号码', '手机号', '电话') or '').strip() or None
                department = sanitize((pick(row, 'department', '所在单位', '单位') or '').strip()) or None

                if not name or not id_card:
                    error_count += 1
                    continue
                if not _is_valid_cn_id_card(id_card):
                    error_count += 1
                    continue

                full_py, initials_py = _name_pinyin_parts(name)
                age = _age_from_id_card(id_card)

                existing = Patient.query.filter_by(id_card=id_card).first()
                if existing:
                    existing.name = name
                    existing.name_pinyin = full_py
                    existing.name_initials = initials_py
                    existing.gender = gender or existing.gender
                    existing.phone = phone if phone else existing.phone
                    existing.department = department if department else existing.department
                    existing.patient_type = 'staff'
                    existing.is_temporary = False
                    existing.age = age
                else:
                    new_patient = Patient(
                        name=name, name_pinyin=full_py, name_initials=initials_py,
                        gender=gender or None, phone=phone, id_card=id_card,
                        department=department, patient_type='staff', is_temporary=False, age=age
                    )
                    db.session.add(new_patient)
                success_count += 1

        elif import_type == 'shop':
            # 商铺员工导入: 姓名,性别,身份证号,手机号码,商铺名称
            for row in csv_input:
                name = sanitize((pick(row, 'name', '姓名') or '').strip())
                gender = (pick(row, 'gender', '性别') or '').strip()
                id_card = (pick(row, 'id_card', '身份证号') or '').strip()
                phone = (pick(row, 'phone', '手机号码', '手机号', '电话') or '').strip() or None
                shop_name = sanitize((pick(row, 'shop_name', '商铺名称', '商铺') or '').strip()) or None

                if not name or not id_card:
                    error_count += 1
                    continue
                if not _is_valid_cn_id_card(id_card):
                    error_count += 1
                    continue

                full_py, initials_py = _name_pinyin_parts(name)
                age = _age_from_id_card(id_card)

                existing = Patient.query.filter_by(id_card=id_card).first()
                if existing:
                    existing.name = name
                    existing.name_pinyin = full_py
                    existing.name_initials = initials_py
                    existing.gender = gender or existing.gender
                    existing.phone = phone if phone else existing.phone
                    existing.shop_name = shop_name if shop_name else existing.shop_name
                    existing.patient_type = 'shop'
                    existing.is_temporary = False
                    existing.age = age
                else:
                    new_patient = Patient(
                        name=name, name_pinyin=full_py, name_initials=initials_py,
                        gender=gender or None, phone=phone, id_card=id_card,
                        shop_name=shop_name, patient_type='shop', is_temporary=False, age=age
                    )
                    db.session.add(new_patient)
                success_count += 1

        else:
            # 学生导入（保持原有逻辑）
            for row in csv_input:
                student_id = sanitize((pick(row, 'student_id', '学号') or '').strip())
                name = sanitize((pick(row, 'name', '姓名') or '').strip())
                gender = (pick(row, 'gender', '性别') or '').strip()
                phone_raw = (pick(row, 'phone', '手机号码', '手机号', '电话') or '').strip()
                grade = (pick(row, 'grade', '年级') or '').strip()
                college = (pick(row, 'college', '学院') or '').strip()
                major = (pick(row, 'major', '专业') or '').strip()
                class_name = (pick(row, 'class_name', '班级') or '').strip()
                age_raw = (pick(row, 'age', '年龄') or '').strip()
                counselor_name = sanitize((pick(row, 'counselor_name', '辅导员姓名', '辅导员') or '').strip())
                phone = phone_raw or None

                missing = []
                if not student_id:
                    missing.append("student_id")
                if not name:
                    missing.append("name")
                if not class_name:
                    missing.append("class_name")
                if not counselor_name:
                    missing.append("counselor_name")

                if missing:
                    error_count += 1
                    continue

                full_py, initials_py = _name_pinyin_parts(name)

                age_val = None
                if age_raw:
                    try:
                        age_val = int(age_raw)
                    except ValueError:
                        error_count += 1
                        continue

                existing = Patient.query.filter_by(student_id=student_id).first()
                if existing:
                    existing.name = name
                    existing.name_pinyin = full_py
                    existing.name_initials = initials_py
                    existing.patient_type = 'student'
                    existing.is_temporary = False
                    if gender:
                        existing.gender = gender
                    if grade:
                        existing.grade = grade
                    if college:
                        existing.college = college
                    if major:
                        existing.major = major
                    existing.class_name = class_name
                    existing.counselor_name = counselor_name
                    if age_raw:
                        existing.age = age_val
                    if phone is not None:
                        existing.phone = phone
                else:
                    new_patient = Patient(
                        student_id=student_id, name=name, name_pinyin=full_py, name_initials=initials_py,
                        gender=gender or None, grade=grade or None, college=college or None,
                        major=major or None, class_name=class_name, phone=phone,
                        age=age_val, counselor_name=counselor_name,
                        patient_type='student', is_temporary=False
                    )
                    db.session.add(new_patient)

                success_count += 1

        type_label = PATIENT_TYPES.get(import_type, '人员')
        log = OperationLog(
            user_id=int(get_jwt_identity()),
            action_type='import_data',
            target_type='patient',
            target_id=0,
            summary=f"批量导入{type_label}: 成功{success_count}条, 失败{error_count}条"
        )
        db.session.add(log)
        db.session.commit()
        return jsonify({"msg": f"Import complete. Success: {success_count}, Errors: {error_count}"}), 200

    except Exception:
        db.session.rollback()
        current_app.logger.exception("Patient CSV import failed")
        return jsonify({"msg": "CSV 导入失败，请检查文件格式或联系管理员"}), 500

@bp.route('/admin/patients', methods=['GET'])
@role_required('admin')
def get_patients():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('size', 20, type=int)
    keyword = request.args.get('keyword', '')
    patient_type = request.args.get('patient_type', '')

    query = Patient.query.order_by(Patient.id.desc())
    if patient_type:
        query = query.filter(Patient.patient_type == patient_type)
    if keyword:
        like = f'%{keyword}%'
        query = query.filter(
            db.or_(
                Patient.name.like(like),
                Patient.student_id.like(like),
                Patient.phone.like(like),
                Patient.name_pinyin.like(like),
                Patient.name_initials.like(like),
                Patient.id_card.like(like),
            )
        )

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    data = []
    for p in pagination.items:
        data.append({
            "id": p.id,
            "student_id": p.student_id,
            "name": p.name,
            "gender": p.gender,
            "class_name": p.class_name,
            "phone": p.phone,
            "grade": p.grade,
            "college": p.college,
            "major": p.major,
            "age": p.age,
            "id_card": p.id_card,
            "counselor_name": p.counselor_name,
            "is_temporary": p.is_temporary,
            "patient_type": p.patient_type or 'student',
            "department": p.department,
            "shop_name": p.shop_name,
        })

    return jsonify({
        "data": data,
        "meta": {"page": page, "per_page": per_page, "total": pagination.total}
    }), 200

@bp.route('/admin/patients', methods=['POST'])
@role_required('admin')
def admin_create_patient():
    data = request.get_json() or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({"msg": "姓名不能为空"}), 400

    patient_type = (data.get('patient_type') or 'student').strip()
    is_temporary = (patient_type == 'temporary')
    # 兼容旧接口：若传了 is_temporary=true 且无 patient_type，则自动设为 temporary
    if 'patient_type' not in data and data.get('is_temporary'):
        patient_type = 'temporary'
        is_temporary = True

    full_py, initials_py = _name_pinyin_parts(name)

    # 通用字段
    gender = (data.get('gender') or '').strip() or None
    phone = (data.get('phone') or '').strip() or None
    id_card = (data.get('id_card') or '').strip() or None
    try:
        age = _nonnegative_int(data['age'], 'age') if data.get('age') else None
    except ValueError:
        return jsonify({"msg": "年龄必须为非负整数", "field": "age"}), 400
    if patient_type not in PATIENT_TYPES:
        return jsonify({"msg": "人员类型不合法", "field": "patient_type"}), 400

    # 类型专属字段初始化
    student_id = None
    class_name = None
    grade = None
    college = None
    major = None
    counselor_name = None
    department = None
    shop_name = None

    if patient_type == 'student':
        student_id = (data.get('student_id') or '').strip() or None
        if student_id:
            existing = Patient.query.filter_by(student_id=student_id).first()
            if existing:
                return jsonify({"msg": f"学号 {student_id} 已存在"}), 400
        class_name = (data.get('class_name') or '').strip() or None
        grade = (data.get('grade') or '').strip() or None
        college = (data.get('college') or '').strip() or None
        major = (data.get('major') or '').strip() or None
        counselor_name = (data.get('counselor_name') or '').strip() or None

    elif patient_type == 'staff':
        if not id_card:
            return jsonify({"msg": "教职工必须填写身份证号"}), 400
        if not _is_valid_cn_id_card(id_card):
            return jsonify({"msg": "身份证号格式不正确"}), 400
        age = _age_from_id_card(id_card)
        department = (data.get('department') or '').strip() or None

    elif patient_type == 'shop':
        if not id_card:
            return jsonify({"msg": "商铺员工必须填写身份证号"}), 400
        if not _is_valid_cn_id_card(id_card):
            return jsonify({"msg": "身份证号格式不正确"}), 400
        age = _age_from_id_card(id_card)
        shop_name = (data.get('shop_name') or '').strip() or None

    elif patient_type == 'temporary':
        if not phone:
            return jsonify({"msg": "临时人员必须填写手机号"}), 400
        if id_card and not _is_valid_cn_id_card(id_card):
            return jsonify({"msg": "身份证号格式不正确"}), 400

    patient = Patient(
        student_id=student_id, name=name, name_pinyin=full_py, name_initials=initials_py,
        gender=gender, class_name=class_name, phone=phone, grade=grade,
        college=college, major=major, age=age, id_card=id_card,
        counselor_name=counselor_name, is_temporary=is_temporary,
        patient_type=patient_type, department=department, shop_name=shop_name,
    )
    db.session.add(patient)
    db.session.flush()

    type_label = PATIENT_TYPES.get(patient_type, '人员')
    log = OperationLog(
        user_id=int(get_jwt_identity()),
        action_type='create_patient',
        target_type='patient',
        target_id=patient.id,
        summary=f'新增{type_label}: {name}'
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({"data": {"id": patient.id}}), 201

@bp.route('/admin/patients/<int:id>', methods=['PUT'])
@role_required('admin')
def admin_update_patient(id):
    patient = db.get_or_404(Patient, id)
    data = request.get_json() or {}

    # 类型变更处理
    if 'patient_type' in data:
        new_type = (data['patient_type'] or '').strip()
        if new_type and new_type != patient.patient_type:
            patient.patient_type = new_type
            patient.is_temporary = (new_type == 'temporary')
            # 切换到非学生类型时清空学生专属字段
            if new_type != 'student':
                for sf in ['student_id', 'class_name', 'grade', 'college', 'major', 'counselor_name']:
                    setattr(patient, sf, None)
            # 切换到非教职工类型时清空 department
            if new_type != 'staff':
                patient.department = None
            # 切换到非商铺类型时清空 shop_name
            if new_type != 'shop':
                patient.shop_name = None

    # 兼容旧接口 is_temporary 参数
    if 'is_temporary' in data and 'patient_type' not in data:
        patient.is_temporary = bool(data['is_temporary'])
        if data['is_temporary']:
            patient.patient_type = 'temporary'

    new_student_id = (data.get('student_id') or '').strip() or None
    if new_student_id and new_student_id != patient.student_id:
        existing = Patient.query.filter_by(student_id=new_student_id).first()
        if existing:
            return jsonify({"msg": f"学号 {new_student_id} 已存在"}), 400
        patient.student_id = new_student_id

    if 'name' in data:
        name = (data['name'] or '').strip()
        if not name:
            return jsonify({"msg": "姓名不能为空"}), 400
        patient.name = name
        full_py, initials_py = _name_pinyin_parts(name)
        patient.name_pinyin = full_py
        patient.name_initials = initials_py

    for field in ['gender', 'class_name', 'phone', 'grade', 'college', 'major', 'id_card', 'counselor_name', 'department', 'shop_name']:
        if field in data:
            setattr(patient, field, (data[field] or '').strip() or None)

    # 教职工/商铺员工：id_card 更新时自动重算 age
    if patient.patient_type in ('staff', 'shop') and patient.id_card:
        new_age = _age_from_id_card(patient.id_card)
        if new_age is not None:
            patient.age = new_age
    elif 'age' in data:
        patient.age = int(data['age']) if data['age'] else None

    log = OperationLog(
        user_id=int(get_jwt_identity()),
        action_type='update_patient',
        target_type='patient',
        target_id=patient.id,
        summary=f"编辑人员: {patient.name}"
    )
    db.session.add(log)
    db.session.commit()
    return jsonify({"msg": "更新成功"}), 200

@bp.route('/admin/patients/<int:id>', methods=['DELETE'])
@role_required('admin')
def delete_patient(id):
    patient = db.get_or_404(Patient, id)
    if patient.visits.count() > 0:
        return jsonify({"msg": "该人员已有就诊记录，无法删除"}), 400

    db.session.delete(patient)
    db.session.commit()
    return jsonify({"msg": "删除成功"}), 200

@bp.route('/admin/patients/<int:id>/visits', methods=['GET'])
@role_required('admin')
def get_patient_visits(id):
    patient = db.get_or_404(Patient, id)
    visits = patient.visits.order_by(Visit.timestamp.desc()).all()

    data = []
    for v in visits:
        doctor_name = v.doctor.real_name if v.doctor else '-'
        data.append({
            "visit_id": v.id,
            "date": format_local_datetime(v.timestamp),
            "diagnosis": v.diagnosis,
            "chief_complaint": v.chief_complaint,
            "total_amount": v.total_amount,
            "status": v.status,
            "doctor_name": doctor_name,
        })

    return jsonify({"data": data}), 200

@bp.route('/admin/visits/<int:visit_id>', methods=['GET'])
@role_required('admin')
def admin_get_visit_detail(visit_id):
    visit = db.get_or_404(Visit, visit_id)

    items = []
    for item in visit.items:
        items.append({
            "drug_name": item.drug.name if item.drug else '-',
            "specification": item.drug.specification if item.drug else '-',
            "usage": item.usage,
            "dosage": item.dosage,
            "frequency": item.frequency,
            "timing": item.timing,
            "days": item.days,
            "quantity": item.quantity,
            "price_at_visit": item.price_at_visit,
            "amount": item.amount,
        })

    return jsonify({
        "data": {
            "id": visit.id,
            "created_at": format_local_datetime(visit.timestamp),
            "status": visit.status,
            "chief_complaint": visit.chief_complaint,
            "present_illness": visit.present_illness,
            "past_history": visit.past_history,
            "physical_exam": visit.physical_exam,
            "diagnosis": visit.diagnosis,
            "doctor_advice": visit.doctor_advice,
            "special_note": visit.special_note,
            "total_amount": visit.total_amount,
            "consultation_fee": visit.consultation_fee,
            "doctor_name": visit.doctor.real_name if visit.doctor else '-',
            "patient": {
                "name": visit.patient.name if visit.patient else '-',
                "student_id": visit.patient.student_id if visit.patient else '-',
                "gender": visit.patient.gender if visit.patient else '-',
            },
            "items": items,
        }
    }), 200

@bp.route('/admin/drugs', methods=['GET'])
@role_required(['admin', 'nurse', 'finance'])
def get_drugs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('size', 20, type=int)
    keyword = request.args.get('keyword', '')

    query = Drug.query.order_by(*nulls_last_asc(Drug.storage_location))
    if keyword:
        query = query.filter(Drug.name.contains(keyword))

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    data = []
    for drug in pagination.items:
        data.append({
            "id": drug.id,
            "name": drug.name,
            "base_name": drug.base_name,
            "type": drug.type,
            "specification": drug.specification,
            "unit": drug.unit,
            "purchase_price": drug.purchase_price,
            "price": drug.price,
            "has_scattered": drug.has_scattered,
            "scattered_price": drug.scattered_price,
            "conversion_rate": drug.conversion_rate,
            "stock": drug.stock,
            "status": drug.status,
            "batch_no": drug.batch_no,
            "inbound_at": format_local_datetime(drug.inbound_at),
            "variant_type": drug.variant_type,
            "stock_group_code": drug.stock_group_code,
            "unit_amount": drug.unit_amount,
            "storage_location": drug.storage_location,
            "expiry_date": drug.expiry_date.isoformat() if drug.expiry_date else None
        })

    return jsonify({
        "data": data,
        "meta": {
            "page": page,
            "per_page": per_page,
            "total": pagination.total
        }
    }), 200

@bp.route('/admin/drugs', methods=['POST'])
@role_required(['admin', 'nurse'])
@serialized_stock_mutation
def create_drug():
    data = request.get_json() or {}

    try:
        drug_type = int(data.get('type', 1))
    except (TypeError, ValueError):
        return jsonify({"msg": "Invalid type", "field": "type"}), 400
    if drug_type not in (1, 2, 3):
        return jsonify({"msg": "Invalid type", "field": "type"}), 400

    if drug_type == 1 or drug_type == 3:
        required_fields = ['name', 'specification', 'unit', 'price', 'stock']
    else:
        required_fields = ['name', 'specification', 'unit', 'price']
        data['stock'] = -1

    for field in required_fields:
        if field not in data:
            return jsonify({"msg": f"Missing required field: {field}"}), 400

    if not all(str(data.get(field) or '').strip() for field in ('name', 'specification', 'unit')):
        return jsonify({"msg": "名称、规格和单位不能为空"}), 400

    try:
        price = _nonnegative_float(data['price'], 'price')
        purchase_price = _nonnegative_float(data.get('purchase_price', 0), 'purchase_price')
        stock = _nonnegative_int(data.get('stock', 0), 'stock') if drug_type in (1, 3) else -1
        scattered_price = None
        conversion_rate = None
        if data.get('has_scattered') and drug_type == 1:
            scattered_price = _nonnegative_float(data.get('scattered_price'), 'scattered_price')
            conversion_rate = _nonnegative_int(data.get('conversion_rate'), 'conversion_rate')
            if scattered_price <= 0 or conversion_rate <= 0:
                raise ValueError('scattered_price')
    except ValueError as exc:
        return jsonify({"msg": "价格、库存或拆零参数不合法", "field": str(exc)}), 400

    # 有效期验证
    expiry_val = None
    if data.get('expiry_date'):
        try:
            expiry_val = date.fromisoformat(data['expiry_date'])
            if expiry_val < local_now().date():
                return jsonify({"msg": "有效期不能早于当前日期"}), 400
        except ValueError:
            return jsonify({"msg": "有效期格式错误，请使用 YYYY-MM-DD 格式"}), 400

    try:
        inbound_at = parse_local_datetime(data.get('inbound_at')) if data.get('inbound_at') else None
    except ValueError:
        return jsonify({"msg": "入库时间格式错误"}), 400

    drug = Drug(
        name=data['name'],
        base_name=data['name'],
        type=drug_type,
        specification=data['specification'],
        unit=data['unit'],
        purchase_price=purchase_price,
        price=price,
        has_scattered=data.get('has_scattered', False) if drug_type != 3 else False,
        scattered_price=scattered_price,
        conversion_rate=conversion_rate,
        stock=stock,
        status=data.get('status', 1),
        batch_no=data.get('batch_no'),
        inbound_at=inbound_at,
        variant_type="consumable" if drug_type == 3 else None,
        storage_location=data.get('storage_location'),
        expiry_date=expiry_val,
    )
    db.session.add(drug)
    db.session.flush()  # 获取 drug.id

    # 记录初始库存盘点记录
    if drug_type in (1, 3) and int(data.get('stock', 0)) > 0:
        ir = InventoryRecord(
            drug_id=drug.id,
            nurse_id=int(get_jwt_identity()),
            old_stock=0,
            new_stock=int(data['stock']),
            operation_type=INVENTORY_OPERATION_INBOUND,
            remark='初始入库(管理员新增)',
        )
        db.session.add(ir)

    # 记录运营日志
    log = OperationLog(
        user_id=int(get_jwt_identity()),
        action_type='drug_create',
        target_type='drug',
        target_id=drug.id,
        summary=f"新增{'药品' if drug_type == 1 else '诊疗项目' if drug_type == 2 else '耗材'}: {drug.name}",
        details=json.dumps({
            "name": drug.name,
            "specification": drug.specification or "",
            "unit": drug.unit or "",
            "price": drug.price,
            "stock": drug.stock
        }, ensure_ascii=False)
    )
    db.session.add(log)
    db.session.commit()

    return jsonify({"data": {"id": drug.id}}), 201

@bp.route('/admin/drugs/<int:id>', methods=['PUT'])
@role_required(['admin', 'nurse'])
@serialized_stock_mutation
def update_drug(id):
    drug = Drug.query.filter_by(id=id).with_for_update().first_or_404()
    data = request.get_json() or {}

    if drug.stock_group_code:
        allowed_group_fields = {'id', 'price', 'purchase_price', 'storage_location'}
        unsupported = sorted(set(data) - allowed_group_fields)
        if unsupported:
            return jsonify({
                "msg": "库存组物资只能单独修改售价、购进价和库位",
                "fields": unsupported,
            }), 400
        if 'purchase_price' in data and drug.variant_type == 'retail':
            return jsonify({"msg": "整散库存组购进价请在整装记录上修改"}), 400

    try:
        if 'type' in data:
            data['type'] = int(data['type'])
            if data['type'] not in (1, 2, 3):
                raise ValueError('type')
        if 'purchase_price' in data:
            data['purchase_price'] = _nonnegative_float(data['purchase_price'], 'purchase_price')
        if 'price' in data:
            data['price'] = _nonnegative_float(data['price'], 'price')
        if 'stock' in data:
            data['stock'] = _nonnegative_int(data['stock'], 'stock')
        if data.get('scattered_price') not in (None, ''):
            data['scattered_price'] = _nonnegative_float(data['scattered_price'], 'scattered_price')
        if data.get('conversion_rate') not in (None, ''):
            data['conversion_rate'] = _nonnegative_int(data['conversion_rate'], 'conversion_rate')
            if data['conversion_rate'] <= 0:
                raise ValueError('conversion_rate')
    except (TypeError, ValueError) as exc:
        return jsonify({"msg": "价格、库存或类型参数不合法", "field": str(exc)}), 400

    old_stock = drug.stock
    if 'name' in data: drug.name = data['name']
    if 'type' in data:
        drug.type = data['type']
        # 类型变更时同步更新variant_type和相关字段
        if data['type'] == 3:
            drug.variant_type = 'consumable'
            drug.has_scattered = False
            drug.scattered_price = None
            drug.conversion_rate = None
        elif data['type'] == 2:
            drug.variant_type = 'service'
            drug.has_scattered = False
            drug.scattered_price = None
            drug.conversion_rate = None
        elif data['type'] == 1 and drug.variant_type in ('service', 'consumable'):
            drug.variant_type = None
    if 'specification' in data: drug.specification = data['specification']
    if 'unit' in data: drug.unit = data['unit']
    if 'purchase_price' in data: drug.purchase_price = data['purchase_price']
    if 'price' in data: drug.price = data['price']
    if 'has_scattered' in data: drug.has_scattered = data['has_scattered']
    if 'scattered_price' in data: drug.scattered_price = data['scattered_price'] if data['scattered_price'] not in (None, '') else None
    if 'conversion_rate' in data: drug.conversion_rate = data['conversion_rate'] if data['conversion_rate'] not in (None, '') else None
    if 'stock' in data: drug.stock = data['stock']
    if 'status' in data: drug.status = int(data['status'])
    if 'batch_no' in data: drug.batch_no = data['batch_no'] or None
    if 'inbound_at' in data:
        try:
            drug.inbound_at = parse_local_datetime(data['inbound_at']) if data['inbound_at'] else None
        except ValueError:
            return jsonify({"msg": "入库时间格式错误"}), 400
    if 'storage_location' in data: drug.storage_location = data['storage_location'] or None
    if 'expiry_date' in data:
        if data['expiry_date']:
            try:
                new_expiry = date.fromisoformat(data['expiry_date'])
                if new_expiry < local_now().date():
                    return jsonify({"msg": "有效期不能早于当前日期"}), 400
                drug.expiry_date = new_expiry
            except ValueError:
                return jsonify({"msg": "有效期格式错误，请使用 YYYY-MM-DD 格式"}), 400
        else:
            drug.expiry_date = None

    if drug.stock_group_code and 'purchase_price' in data:
        group = DrugStockGroup.query.filter_by(
            group_code=drug.stock_group_code
        ).with_for_update().populate_existing().first()
        if group and group.retail_drug and group.retail_amount:
            group.retail_drug.purchase_price = round(
                float(data['purchase_price']) * float(group.retail_amount) / float(group.pack_amount),
                2,
            )

    # 若库存发生变化，记录盘点记录
    if 'stock' in data and drug.type in (1, 3) and int(data['stock']) != old_stock:
        ir = InventoryRecord(
            drug_id=drug.id,
            nurse_id=int(get_jwt_identity()),
            old_stock=old_stock,
            new_stock=int(data['stock']),
            operation_type=INVENTORY_OPERATION_ADJUSTMENT,
            remark='管理员编辑库存',
        )
        db.session.add(ir)

    log = OperationLog(
        user_id=int(get_jwt_identity()),
        action_type='drug_update',
        target_type='drug',
        target_id=drug.id,
        summary=f"编辑{'药品' if drug.type == 1 else '诊疗项目' if drug.type == 2 else '耗材'}: {drug.name}"
    )
    db.session.add(log)
    db.session.commit()
    return jsonify({"msg": "Drug updated successfully"}), 200

@bp.route('/admin/drugs/<int:drug_id>/inbound', methods=['POST'])
@role_required(['admin', 'nurse'])
@serialized_stock_mutation
def drug_inbound(drug_id):
    """药品入库（进货/补货）"""
    drug = Drug.query.filter_by(id=drug_id).with_for_update().first_or_404()
    data = request.get_json() or {}

    try:
        quantity = _nonnegative_int(data.get('quantity'), 'quantity')
    except ValueError:
        return jsonify({"msg": "入库数量必须为正整数", "field": "quantity"}), 400
    remark = data.get('remark', '')

    if quantity <= 0:
        return jsonify({"msg": "入库数量必须大于0"}), 400
    if drug.type not in (1, 3) and drug.type is not None:
        return jsonify({"msg": "诊疗项目不支持库存入库"}), 400
    if drug.stock_group_code:
        return jsonify({"msg": "库存组物资请使用护士入库流程补货"}), 400

    old_stock = int(drug.stock or 0)
    drug.stock = old_stock + quantity

    # 创建入库记录
    record = InventoryRecord(
        drug_id=drug.id,
        nurse_id=int(get_jwt_identity()),
        old_stock=old_stock,
        new_stock=drug.stock,
        operation_type=INVENTORY_OPERATION_INBOUND,
        remark=f"入库: {remark}" if remark else "入库"
    )
    db.session.add(record)
    db.session.commit()

    return jsonify({
        "msg": "入库成功",
        "data": {
            "id": drug.id,
            "name": drug.name,
            "old_stock": old_stock,
            "new_stock": drug.stock
        }
    }), 200

@bp.route('/admin/drugs/<int:id>', methods=['DELETE'])
@role_required(['admin', 'nurse'])
@serialized_stock_mutation
def delete_drug(id):
    drug = Drug.query.filter_by(id=id).with_for_update().first_or_404()
    if drug.stock_group_code:
        return jsonify({"msg": "整散库存组药品不支持直接删除，请先停用或由管理员做库存组清理"}), 400
    try:
        db.session.delete(drug)
        db.session.commit()
        return jsonify({"msg": "删除成功"}), 200
    except IntegrityError:
        db.session.rollback()
        return jsonify({"msg": "已被处方引用的项目无法彻底删除，请使用停用功能"}), 400
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Drug deletion failed: id=%s", id)
        return jsonify({"msg": "删除失败，请稍后重试"}), 500

@bp.route('/admin/drugs/import', methods=['POST'])
@role_required(['admin', 'nurse'])
@serialized_stock_mutation
def import_drugs():
    if 'file' not in request.files:
        return jsonify({"msg": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"msg": "No selected file"}), 400
    if not file.filename.endswith('.csv'):
        return jsonify({"msg": "Only CSV files are allowed"}), 400
    if not _check_upload_size(file):
        return jsonify({"msg": "File too large, max 10MB"}), 400

    try:
        file_content = file.stream.read()
        try:
            decoded_content = file_content.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                decoded_content = file_content.decode("gbk")
            except UnicodeDecodeError:
                decoded_content = file_content.decode("utf-8")

        stream = io.StringIO(decoded_content, newline=None)
        csv_input = csv.DictReader(stream)

        success_count = 0
        error_count = 0

        for row in csv_input:
            savepoint = db.session.begin_nested()
            try:
                name = (row.get('name') or '').strip()
                specification = (row.get('specification') or '').strip()
                unit = (row.get('unit') or '').strip()
                purchase_price = _nonnegative_float(row.get('purchase_price') or 0.0, 'purchase_price')
                price = _nonnegative_float(row.get('price') or 0.0, 'price')
                has_scattered = str(row.get('has_scattered', '')).strip() == '1'
                scattered_price = _nonnegative_float(row.get('scattered_price'), 'scattered_price') if row.get('scattered_price') else None
                conversion_rate = _nonnegative_int(row.get('conversion_rate'), 'conversion_rate') if row.get('conversion_rate') else None
                if conversion_rate is not None and conversion_rate <= 0:
                    raise ValueError('conversion_rate')
                batch_no = (row.get('batch_no') or '').strip() or None
                inbound_at = (row.get('inbound_at') or '').strip() or None
                expiry_raw = (row.get('expiry_date') or '').strip()
                expiry_date = date.fromisoformat(expiry_raw) if expiry_raw else None
                if expiry_date and expiry_date < local_now().date():
                    raise ValueError('expiry_date')
                
                stock_str = str(row.get('stock', '')).strip()
                stock = _nonnegative_int(stock_str, 'stock') if stock_str else 0

                if not name or not specification:
                    savepoint.rollback()
                    error_count += 1
                    continue

                existing = Drug.query.filter(
                    Drug.type == 1,
                    Drug.name == name,
                    Drug.specification == specification,
                    Drug.batch_no == batch_no,
                    Drug.expiry_date == expiry_date,
                    Drug.stock_group_code.is_(None),
                    Drug.variant_type.is_(None),
                ).first()
                if existing:
                    old_stock_existing = existing.stock
                    existing.stock += stock
                    existing.purchase_price = purchase_price
                    existing.price = price
                    existing.has_scattered = has_scattered
                    existing.scattered_price = scattered_price
                    existing.conversion_rate = conversion_rate
                    if inbound_at:
                        existing.inbound_at = parse_local_datetime(inbound_at)
                    existing.status = 1
                    # 记录入库盘点记录
                    if stock > 0:
                        db.session.add(InventoryRecord(
                            drug_id=existing.id,
                            nurse_id=int(get_jwt_identity()),
                            old_stock=old_stock_existing,
                            new_stock=existing.stock,
                            operation_type=INVENTORY_OPERATION_INBOUND,
                            remark='CSV批量入库',
                        ))
                else:
                    new_drug = Drug(
                        name=name,
                        specification=specification,
                        unit=unit,
                        purchase_price=purchase_price,
                        price=price,
                        has_scattered=has_scattered,
                        scattered_price=scattered_price,
                        conversion_rate=conversion_rate,
                        stock=stock,
                        status=1,
                        batch_no=batch_no,
                        inbound_at=parse_local_datetime(inbound_at) if inbound_at else None,
                        expiry_date=expiry_date,
                    )
                    db.session.add(new_drug)
                    db.session.flush()
                    # 记录初始入库盘点记录
                    if stock > 0:
                        db.session.add(InventoryRecord(
                            drug_id=new_drug.id,
                            nurse_id=int(get_jwt_identity()),
                            old_stock=0,
                            new_stock=stock,
                            operation_type=INVENTORY_OPERATION_INBOUND,
                            remark='CSV初始入库',
                        ))

                db.session.flush()
                savepoint.commit()
                success_count += 1
            except Exception:
                savepoint.rollback()
                error_count += 1

        db.session.add(OperationLog(
            user_id=int(get_jwt_identity()),
            action_type='import_data',
            target_type='drug',
            target_id=0,
            summary=f'CSV批量导入药品: 成功{success_count}条, 失败{error_count}条',
        ))
        db.session.commit()
        return jsonify({
            "msg": f"Import completed. Success: {success_count}, Failed: {error_count}",
            "data": {"success": success_count, "failed": error_count}
        }), 200

    except Exception:
        db.session.rollback()
        current_app.logger.exception("Drug CSV import failed")
        return jsonify({"msg": "导入失败，请检查文件格式或联系管理员"}), 500

@bp.route('/admin/drugs/import_xls', methods=['POST'])
@role_required(['admin', 'nurse'])
@serialized_stock_mutation
def import_drugs_xls():
    if 'file' not in request.files:
        return jsonify({"msg": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"msg": "No selected file"}), 400
    allowed_extensions = ('.xls', '.xlsx')
    if not file.filename.lower().endswith(allowed_extensions):
        return jsonify({"msg": "Only Excel files (.xls, .xlsx) are allowed"}), 400
    if not _check_upload_size(file):
        return jsonify({"msg": "File too large, max 10MB"}), 400
    if file.filename.lower().endswith('.xlsx'):
        try:
            with zipfile.ZipFile(file.stream) as archive:
                expanded_size = sum(item.file_size for item in archive.infolist())
                if expanded_size > 100 * 1024 * 1024:
                    return jsonify({"msg": "Excel 解压后体积过大"}), 400
        except zipfile.BadZipFile:
            return jsonify({"msg": "Excel 文件格式无效"}), 400
        finally:
            file.stream.seek(0)

    try:
        import pandas as pd
        df = pd.read_excel(file, sheet_name=0)
        if len(df.index) > 10000 or len(df.columns) > 100:
            return jsonify({"msg": "Excel 最多允许 10000 行、100 列"}), 400
        df.columns = [str(col).strip() for col in df.columns]

        success_count = 0
        error_count = 0

        for seq, group in df.groupby('序号'):
            savepoint = db.session.begin_nested()
            try:
                base_name = str(group.iloc[0].get('药  名', '')).strip()
                if not base_name or base_name == 'nan':
                    savepoint.rollback()
                    error_count += 1
                    continue

                spec = str(group.iloc[0].get('规格', '')).strip()
                unit = str(group.iloc[0].get('单位', '')).strip()
                
                # Whole info
                whole_row = group[group['散装价'].isna() | (group['散装价'] == '')]
                if whole_row.empty:
                    whole_row = group.iloc[[0]]
                else:
                    whole_row = whole_row.iloc[[0]]

                purchase_price = _nonnegative_float(whole_row['购进价'].values[0], 'purchase_price') if pd.notnull(whole_row['购进价'].values[0]) else 0.0
                price = _nonnegative_float(whole_row['盒装价'].values[0], 'price') if pd.notnull(whole_row['盒装价'].values[0]) else 0.0
                
                # Scattered info
                scattered_row = group[group['散装价'].notna() & (group['散装价'] != '')]
                has_scattered = not scattered_row.empty

                scattered_price = None
                conversion_rate = None

                if has_scattered:
                    scattered_price = _nonnegative_float(scattered_row['盒装价'].values[0], 'scattered_price')
                    scattered_total = _nonnegative_float(scattered_row['散装价'].values[0], 'scattered_total')
                    if scattered_price > 0:
                        conversion_rate = int(round(scattered_total / scattered_price))
                    else:
                        conversion_rate = 1

                stock = 0
                if '库存' in group.columns:
                    val = group.iloc[0].get('库存')
                    stock = _nonnegative_int(val, 'stock') if pd.notnull(val) else 0

                batch_no = None
                if '批号' in group.columns:
                    val = group.iloc[0].get('批号')
                    batch_no = str(val).strip() if pd.notnull(val) and str(val).strip() else None

                inbound_at = None
                if '入库时间' in group.columns:
                    val = group.iloc[0].get('入库时间')
                    inbound_at = parse_local_datetime(val) if pd.notnull(val) else None

                new_drug = Drug(
                    name=base_name,
                    specification=spec,
                    unit=unit,
                    price=price,
                    purchase_price=purchase_price,
                    has_scattered=has_scattered,
                    scattered_price=scattered_price,
                    conversion_rate=conversion_rate,
                    stock=stock,
                    status=1,
                    batch_no=batch_no,
                    inbound_at=inbound_at
                )
                db.session.add(new_drug)
                db.session.flush()
                # 记录初始入库盘点记录
                if stock > 0:
                    db.session.add(InventoryRecord(
                        drug_id=new_drug.id,
                        nurse_id=int(get_jwt_identity()),
                        old_stock=0,
                        new_stock=stock,
                        operation_type=INVENTORY_OPERATION_INBOUND,
                        remark='Excel批量初始入库',
                    ))
                db.session.flush()
                savepoint.commit()
                success_count += 1
            except Exception:
                savepoint.rollback()
                error_count += 1

        db.session.add(OperationLog(
            user_id=int(get_jwt_identity()),
            action_type='import_data',
            target_type='drug',
            target_id=0,
            summary=f'Excel批量导入药品: 成功{success_count}条, 失败{error_count}条',
        ))
        db.session.commit()
        return jsonify({
            "msg": f"Import completed. Success: {success_count}, Failed: {error_count}",
            "data": {"success": success_count, "failed": error_count}
        }), 200
    except Exception:
        db.session.rollback()
        current_app.logger.exception("Drug Excel import failed")
        return jsonify({"msg": "导入失败，请检查文件格式或联系管理员"}), 500

@bp.route('/admin/drugs/template', methods=['GET'])
@role_required(['admin', 'nurse', 'finance'])
def get_drug_template():
    si = io.BytesIO()
    si.write(b'\xef\xbb\xbf')

    str_io = io.StringIO()
    cw = csv.writer(str_io)
    cw.writerow(['name', 'specification', 'unit', 'purchase_price', 'price', 'has_scattered', 'scattered_price', 'conversion_rate', 'stock', 'batch_no', 'inbound_at'])
    cw.writerow(['示例药品', '10mg*10片', '盒', '8.5', '10.5', '1', '1.2', '10', '100', 'BATCH-001', '2026-03-01 10:30'])
    cw.writerow(['无库存无零卖示例', '100ml', '瓶', '12', '15', '0', '', '', '0', '', ''])

    si.write(str_io.getvalue().encode('utf-8'))

    output = make_response(si.getvalue())
    output.headers["Content-Disposition"] = "attachment; filename=drug_import_template.csv"
    output.headers["Content-type"] = "text/csv; charset=utf-8"
    return output

@bp.route('/admin/drugs/smart-inventory', methods=['POST'])
@role_required(['admin', 'nurse'])
@serialized_stock_mutation
def smart_inventory():
    data = request.get_json() or {}
    try:
        threshold = int(data.get('threshold', 30))
        expiry_threshold = int(data.get('expiry_threshold', 30))
    except (TypeError, ValueError):
        return jsonify({"msg": "预警阈值必须为整数"}), 400
    if threshold < 0 or threshold > 100000:
        return jsonify({"msg": "库存预警阈值超出范围"}), 400
    if expiry_threshold < 1 or expiry_threshold > 3650:
        return jsonify({"msg": "有效期预警天数超出范围"}), 400
    scattered_only = data.get('scattered_only', False)
    confirm_merge = data.get('confirm_merge') is True
    requested_candidate_ids = []
    if confirm_merge:
        raw_candidates = data.get('merge_candidate_ids')
        if not isinstance(raw_candidates, list) or not raw_candidates:
            return jsonify({"msg": "确认合并时必须提交预览候选记录"}), 400
        used_ids = set()
        try:
            for raw_group in raw_candidates:
                if not isinstance(raw_group, list) or len(raw_group) < 2:
                    raise ValueError
                group_ids = tuple(sorted(int(value) for value in raw_group))
                if any(value <= 0 for value in group_ids) or len(set(group_ids)) != len(group_ids):
                    raise ValueError
                if used_ids.intersection(group_ids):
                    raise ValueError
                used_ids.update(group_ids)
                requested_candidate_ids.append(group_ids)
        except (TypeError, ValueError):
            return jsonify({"msg": "合并候选记录格式不合法"}), 400

    total_merged = 0
    total_deleted = 0

    try:
        merge_fields = (
            Drug.type,
            Drug.name,
            Drug.specification,
            Drug.unit,
            Drug.batch_no,
            Drug.expiry_date,
            Drug.price,
            Drug.purchase_price,
            Drug.variant_type,
            Drug.has_scattered,
            Drug.scattered_price,
            Drug.conversion_rate,
            Drug.storage_location,
            Drug.status,
            Drug.inbound_at,
        )
        duplicates_query = db.session.query(
            *merge_fields,
            func.count(Drug.id).label('count')
        ).filter(
            Drug.stock_group_code.is_(None)
        ).group_by(*merge_fields).having(func.count(Drug.id) > 1).all()

        merge_candidates = []
        candidate_groups = []

        for dup in duplicates_query:
            filters = [
                Drug.stock_group_code.is_(None),
                Drug.type == dup.type,
                Drug.name == dup.name,
                Drug.specification == dup.specification,
                Drug.unit == dup.unit,
                Drug.batch_no == dup.batch_no,
                Drug.expiry_date == dup.expiry_date,
                Drug.price == dup.price,
                Drug.purchase_price == dup.purchase_price,
                Drug.variant_type == dup.variant_type,
                Drug.has_scattered == dup.has_scattered,
                Drug.scattered_price == dup.scattered_price,
                Drug.conversion_rate == dup.conversion_rate,
                Drug.storage_location == dup.storage_location,
                Drug.status == dup.status,
                Drug.inbound_at == dup.inbound_at,
            ]
            drugs = Drug.query.filter(*filters).order_by(Drug.id.asc()).all()
            if len(drugs) <= 1:
                continue

            record_ids = tuple(drug.id for drug in drugs)
            candidate_groups.append((record_ids, filters))
            merge_candidates.append({
                "name": dup.name,
                "specification": dup.specification,
                "batch_no": dup.batch_no,
                "expiry_date": dup.expiry_date.isoformat() if dup.expiry_date else None,
                "record_ids": list(record_ids),
                "record_count": len(drugs),
                "combined_stock": sum(int(drug.stock or 0) for drug in drugs),
            })

        if confirm_merge:
            candidate_map = {record_ids: filters for record_ids, filters in candidate_groups}
            if any(record_ids not in candidate_map for record_ids in requested_candidate_ids):
                db.session.rollback()
                return jsonify({
                    "msg": "合并候选已发生变化，请重新预览后确认",
                    "data": {"merge_candidates": merge_candidates},
                }), 409

            selected_candidates = []
            for record_ids in sorted(requested_candidate_ids):
                filters = candidate_map[record_ids]
                drugs = (
                    Drug.query.filter(*filters)
                    .order_by(Drug.id.asc())
                    .with_for_update()
                    .all()
                )
                if tuple(drug.id for drug in drugs) != record_ids:
                    db.session.rollback()
                    return jsonify({"msg": "合并候选已发生变化，请重新预览后确认"}), 409

                selected_candidates.append(next(
                    candidate for candidate in merge_candidates
                    if tuple(candidate["record_ids"]) == record_ids
                ))

                primary_drug = drugs[0]
                duplicate_drugs = drugs[1:]
                for dup_drug in duplicate_drugs:
                    if (primary_drug.type in (1, 3) or primary_drug.type is None) and dup_drug.stock:
                        primary_drug.stock = int(primary_drug.stock or 0) + int(dup_drug.stock)

                    PrescriptionItem.query.filter_by(drug_id=dup_drug.id).update(
                        {PrescriptionItem.drug_id: primary_drug.id},
                        synchronize_session=False,
                    )
                    InventoryRecord.query.filter_by(drug_id=dup_drug.id).update(
                        {InventoryRecord.drug_id: primary_drug.id},
                        synchronize_session=False,
                    )

                    for snapshot in DailyStockSnapshot.query.filter_by(drug_id=dup_drug.id).all():
                        primary_snapshot = DailyStockSnapshot.query.filter_by(
                            drug_id=primary_drug.id,
                            date=snapshot.date,
                        ).first()
                        if primary_snapshot:
                            primary_snapshot.stock = int(primary_snapshot.stock or 0) + int(snapshot.stock or 0)
                            db.session.delete(snapshot)
                        else:
                            snapshot.drug_id = primary_drug.id

                    db.session.delete(dup_drug)
                    total_deleted += 1

                if primary_drug.type in (1, 3) or primary_drug.type is None:
                    merged_stock = int(primary_drug.stock or 0)
                    db.session.add(InventoryRecord(
                        drug_id=primary_drug.id,
                        nurse_id=int(get_jwt_identity()),
                        old_stock=merged_stock,
                        new_stock=merged_stock,
                        operation_type=INVENTORY_OPERATION_MERGE,
                        remark='智能盘库确认合并重复记录',
                    ))
                total_merged += 1

            if total_merged:
                db.session.add(OperationLog(
                    user_id=int(get_jwt_identity()),
                    action_type='smart_inventory_merge',
                    target_type='drug',
                    summary=f'智能盘库确认合并 {total_merged} 组重复记录',
                    details=json.dumps({
                        "merged_groups": total_merged,
                        "deleted_records": total_deleted,
                        "candidates": selected_candidates,
                    }, ensure_ascii=False),
                ))
            db.session.commit()

        # 库存预警药品
        query = Drug.query.filter(
            Drug.type.in_([1, 3]),
            Drug.status == 1,
            Drug.stock < threshold
        )
        if scattered_only:
            query = query.filter(
                db.or_(
                    Drug.name.like('%散%'),
                    Drug.specification.like('%散%')
                )
            )
        low_stock_drugs = query.order_by(Drug.stock.asc()).all()
        warnings = [{"id": d.id, "name": d.name, "specification": d.specification, "stock": d.stock} for d in low_stock_drugs]
        
        # 有效期预警药品（阈值天数内到期）
        today = local_now().date()
        warn_date = today + timedelta(days=expiry_threshold)
        expiry_query = Drug.query.filter(
            Drug.type.in_([1, 3]),
            Drug.status == 1,
            Drug.expiry_date != None,
            Drug.expiry_date <= warn_date
        )
        expiry_drugs = expiry_query.order_by(Drug.expiry_date.asc()).all()
        expiry_warnings = []
        for d in expiry_drugs:
            days_remaining = (d.expiry_date - today).days
            expiry_warnings.append({
                "id": d.id,
                "name": d.name,
                "specification": d.specification,
                "expiry_date": d.expiry_date.isoformat(),
                "days_remaining": days_remaining,
                "stock": d.stock,
                "is_expired": days_remaining < 0
            })
        
        return jsonify({
            "msg": "盘库完成",
            "data": {
                "merged_groups": total_merged,
                "deleted_duplicates": total_deleted,
                "merge_candidates": merge_candidates,
                "merge_confirmation_required": bool(merge_candidates) and not confirm_merge,
                "warnings": warnings,
                "threshold": threshold,
                "expiry_warnings": expiry_warnings,
                "expiry_threshold": expiry_threshold
            }
        }), 200

    except Exception:
        db.session.rollback()
        current_app.logger.exception("Smart inventory failed")
        return jsonify({"msg": "智能盘库失败，请稍后重试"}), 500

@bp.route('/admin/statistics/revenue', methods=['GET'])
@role_required(['admin', 'nurse', 'finance'])
def get_revenue_stats():
    def parse_dt(value, is_end=False):
        return parse_local_datetime(value, is_end=is_end)

    stats_type = request.args.get("type", "daily")
    target_date_str = request.args.get("date", local_now().strftime("%Y-%m-%d"))
    start_time_str = request.args.get("start_time") or request.args.get("start_date")
    end_time_str = request.args.get("end_time") or request.args.get("end_date")

    doctor_id = request.args.get("doctor_id", type=int)
    nurse_id = request.args.get("nurse_id", type=int)

    # 分页参数（仅对明细数据生效，汇总统计始终全量计算）
    try:
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 20))
    except (TypeError, ValueError):
        page = 1
        per_page = 20
    page = max(page, 1)
    per_page = max(min(per_page, 200), 1)  # 限制1-200条/页

    try:
        if start_time_str or end_time_str:
            start = parse_dt(start_time_str, is_end=False) if start_time_str else None
            end = parse_dt(end_time_str, is_end=True) if end_time_str else None
            if start is None and end is None:
                raise ValueError("Invalid date format")
            if start is None:
                start = end - timedelta(days=1)
            if end is None:
                end = start + timedelta(days=1)
        else:
            if stats_type == "daily":
                target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
                start = local_naive_to_utc(datetime.combine(target_date, datetime.min.time()))
                end = start + timedelta(days=1)
            elif stats_type == "monthly":
                target_year, target_month = map(int, target_date_str.split("-")[:2])
                start = local_naive_to_utc(datetime(target_year, target_month, 1))
                if target_month == 12:
                    end = local_naive_to_utc(datetime(target_year + 1, 1, 1))
                else:
                    end = local_naive_to_utc(datetime(target_year, target_month + 1, 1))
            else:
                target_year = int(target_date_str.split("-")[0])
                start = local_naive_to_utc(datetime(target_year, 1, 1))
                end = local_naive_to_utc(datetime(target_year + 1, 1, 1))

        query = Payment.query.join(Visit).filter(
            Payment.payment_date >= start,
            Payment.payment_date < end
        )
        if doctor_id:
            query = query.filter(Visit.doctor_id == doctor_id)
        if nurse_id:
            query = query.filter(Payment.nurse_id == nurse_id)

        payments = query.options(
            db.joinedload(Payment.visit).joinedload(Visit.patient),
            db.joinedload(Payment.visit).joinedload(Visit.doctor),
            db.joinedload(Payment.nurse),
        ).all()

        visit_ids = [p.visit_id for p in payments if p.visit_id]
        items_by_visit = {}
        if visit_ids:
            items = PrescriptionItem.query.options(
                db.joinedload(PrescriptionItem.drug)
            ).filter(
                PrescriptionItem.visit_id.in_(visit_ids)
            ).all()
            for it in items:
                items_by_visit.setdefault(it.visit_id, []).append(it)

        total_revenue = 0.0
        drug_revenue = 0.0
        service_revenue = 0.0
        consumable_revenue = 0.0
        consultation_revenue = 0.0
        total_cost = 0.0
        total_profit = 0.0

        details = []
        finance_view = _current_user_is_finance()
        for p in payments:
            v = p.visit
            if v is None:
                continue

            breakdown = allocate_payment_revenue(p, v, items_by_visit.get(v.id, []))
            consult = breakdown["consultation"]
            drug_amt = breakdown["drug"]
            service_amt = breakdown["service"]
            consumable_amt = breakdown["consumable"]
            amount = breakdown["total"]
            cost = breakdown["cost"]
            profit = breakdown["profit"]

            total_revenue += amount
            drug_revenue += drug_amt
            service_revenue += service_amt
            consumable_revenue += consumable_amt
            consultation_revenue += consult
            total_cost += cost
            total_profit += profit

            details.append({
                "date": format_local_datetime(p.payment_date, "%Y-%m-%d %H:%M:%S") or "",
                "visit_id": p.visit_id,
                "patient_name": _mask_patient_name(v.patient.name if v.patient else "", finance_view),
                "student_id": "" if finance_view else (v.patient.student_id if v.patient else ""),
                "doctor_name": v.doctor.real_name if getattr(v, "doctor", None) else "",
                "nurse_name": p.nurse.real_name if getattr(p, "nurse", None) else "",
                "diagnosis": "" if finance_view else (v.diagnosis or ""),
                "chief_complaint": "" if finance_view else (v.chief_complaint or ""),
                "drug_amount": drug_amt,
                "service_amount": service_amt,
                "consumable_amount": consumable_amt,
                "consultation_fee": consult,
                "amount": amount,
                "cost": cost,
                "profit": profit,
            })

        # 分页切割明细数据（汇总统计仍基于全量数据）
        total = len(details)
        start_idx = (page - 1) * per_page
        end_idx = start_idx + per_page
        paged_details = details[start_idx:end_idx]

        return jsonify({
            "data": {
                "total_revenue": total_revenue,
                "drug_revenue": drug_revenue,
                "service_revenue": service_revenue,
                "consumable_revenue": consumable_revenue,
                "consultation_revenue": consultation_revenue,
                "total_cost": total_cost,
                "total_profit": total_profit,
                "details": paged_details,
                "total": total,
                "page": page,
                "per_page": per_page,
                "range": {
                    "start": utc_naive_to_local(start).strftime("%Y-%m-%d %H:%M:%S"),
                    "end": utc_naive_to_local(end - timedelta(seconds=1)).strftime("%Y-%m-%d %H:%M:%S"),
                }
            }
        }), 200

    except ValueError:
        return jsonify({"msg": "Invalid date format"}), 400


@bp.route('/admin/statistics/revenue/users', methods=['GET'])
@role_required(['admin', 'nurse', 'finance'])
def get_revenue_stats_users():
    users = User.query.filter(User.role.in_(["doctor", "nurse"])).order_by(User.role.asc(), User.real_name.asc()).all()
    doctors = []
    nurses = []
    for u in users:
        item = {"id": u.id, "real_name": u.real_name}
        if u.role == "doctor":
            doctors.append(item)
        elif u.role == "nurse":
            nurses.append(item)
    return jsonify({"data": {"doctors": doctors, "nurses": nurses}}), 200


@bp.route("/admin/statistics/revenue/export", methods=["GET"])
@role_required(['admin', 'nurse', 'finance'])
def export_revenue_stats():
    from openpyxl import Workbook

    stats_type = request.args.get("type", "daily")
    target_date_str = request.args.get("date", local_now().strftime("%Y-%m-%d"))
    start_time_str = request.args.get("start_time") or request.args.get("start_date")
    end_time_str = request.args.get("end_time") or request.args.get("end_date")

    doctor_id = request.args.get("doctor_id", type=int)
    nurse_id = request.args.get("nurse_id", type=int)

    try:
        def parse_dt(value, is_end=False):
            return parse_local_datetime(value, is_end=is_end)

        if start_time_str or end_time_str:
            start = parse_dt(start_time_str, is_end=False) if start_time_str else None
            end = parse_dt(end_time_str, is_end=True) if end_time_str else None
            if start is None and end is None:
                raise ValueError("Invalid date format")
            if start is None:
                start = end - timedelta(days=1)
            if end is None:
                end = start + timedelta(days=1)
            date_label = f"{utc_naive_to_local(start).strftime('%Y%m%d_%H%M')}-{utc_naive_to_local(end - timedelta(seconds=1)).strftime('%Y%m%d_%H%M')}"
        else:
            if stats_type == "daily":
                target_date = datetime.strptime(target_date_str, "%Y-%m-%d").date()
                start = local_naive_to_utc(datetime.combine(target_date, datetime.min.time()))
                end = start + timedelta(days=1)
                date_label = target_date_str
            elif stats_type == "monthly":
                target_year, target_month = map(int, target_date_str.split("-")[:2])
                start = local_naive_to_utc(datetime(target_year, target_month, 1))
                if target_month == 12:
                    end = local_naive_to_utc(datetime(target_year + 1, 1, 1))
                else:
                    end = local_naive_to_utc(datetime(target_year, target_month + 1, 1))
                date_label = f"{target_year:04d}-{target_month:02d}"
            elif stats_type == "yearly":
                target_year = int(target_date_str.split("-")[0])
                start = local_naive_to_utc(datetime(target_year, 1, 1))
                end = local_naive_to_utc(datetime(target_year + 1, 1, 1))
                date_label = f"{target_year:04d}"
            else:
                return jsonify({"msg": "Invalid type"}), 400
    except Exception:
        return jsonify({"msg": "Invalid date format"}), 400

    def safe_text(val):
        s = "" if val is None else str(val)
        if s.startswith(("=", "+", "-", "@")):
            return "'" + s
        return s

    query = Payment.query.join(Visit).filter(Payment.payment_date >= start, Payment.payment_date < end)
    if doctor_id:
        query = query.filter(Visit.doctor_id == doctor_id)
    if nurse_id:
        query = query.filter(Payment.nurse_id == nurse_id)
    payments = query.options(
        db.joinedload(Payment.visit).joinedload(Visit.patient),
        db.joinedload(Payment.visit).joinedload(Visit.doctor),
        db.joinedload(Payment.nurse),
    ).all()

    visit_ids = [p.visit_id for p in payments if p.visit_id]
    items_by_visit = {}
    if visit_ids:
        items = PrescriptionItem.query.options(
            db.joinedload(PrescriptionItem.drug)
        ).filter(
            PrescriptionItem.visit_id.in_(visit_ids)
        ).all()
        for it in items:
            items_by_visit.setdefault(it.visit_id, []).append(it)

    wb = Workbook()
    ws = wb.active
    ws.title = "revenue"

    ws.append(
        [
            "支付时间",
            "visit_id",
            "患者姓名",
            "学号",
            "医生",
            "护士",
            "支付方式",
            "诊察费",
            "药品金额",
            "诊疗项目金额",
            "耗材金额",
            "总金额",
            "成本",
            "利润",
            "诊断",
            "主诉",
        ]
    )

    total_revenue = 0.0
    consultation_revenue = 0.0
    drug_revenue = 0.0
    service_revenue = 0.0
    consumable_revenue = 0.0
    total_cost = 0.0
    total_profit = 0.0

    finance_view = _current_user_is_finance()
    for p in payments:
        v = p.visit
        if v is None:
            continue
        breakdown = allocate_payment_revenue(p, v, items_by_visit.get(v.id, []))
        consult = breakdown["consultation"]
        drug_amt = breakdown["drug"]
        service_amt = breakdown["service"]
        consumable_amt = breakdown["consumable"]
        amount = breakdown["total"]
        cost = breakdown["cost"]
        profit = breakdown["profit"]

        total_revenue += amount
        consultation_revenue += consult
        drug_revenue += drug_amt
        service_revenue += service_amt
        consumable_revenue += consumable_amt
        total_cost += cost
        total_profit += profit

        ws.append(
            [
                safe_text(format_local_datetime(p.payment_date, "%Y-%m-%d %H:%M:%S") or ""),
                v.id,
                safe_text(_mask_patient_name(v.patient.name if v.patient else "", finance_view)),
                safe_text("" if finance_view else (v.patient.student_id if v.patient else "")),
                safe_text(v.doctor.real_name if v.doctor else ""),
                safe_text(p.nurse.real_name if getattr(p, "nurse", None) else ""),
                safe_text(p.payment_method or ""),
                round(consult, 2),
                round(drug_amt, 2),
                round(service_amt, 2),
                round(consumable_amt, 2),
                round(amount, 2),
                round(cost, 2),
                round(profit, 2),
                safe_text("" if finance_view else (v.diagnosis or "")),
                safe_text("" if finance_view else (v.chief_complaint or "")),
            ]
        )

    ws2 = wb.create_sheet("summary")
    ws2.append(["维度", safe_text(stats_type)])
    ws2.append(["日期", safe_text(date_label)])
    ws2.append(["开始时间", safe_text(utc_naive_to_local(start).strftime("%Y-%m-%d %H:%M:%S"))])
    ws2.append(["结束时间", safe_text(utc_naive_to_local(end - timedelta(seconds=1)).strftime("%Y-%m-%d %H:%M:%S"))])
    ws2.append(["doctor_id", safe_text(doctor_id if doctor_id else "")])
    ws2.append(["nurse_id", safe_text(nurse_id if nurse_id else "")])
    ws2.append(["总收入", round(total_revenue, 2)])
    ws2.append(["药品收入", round(drug_revenue, 2)])
    ws2.append(["诊疗项目收入", round(service_revenue, 2)])
    ws2.append(["耗材收入", round(consumable_revenue, 2)])
    ws2.append(["诊察费收入", round(consultation_revenue, 2)])
    ws2.append(["总成本", round(total_cost, 2)])
    ws2.append(["总利润", round(total_profit, 2)])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    filename = f"revenue_{stats_type}_{date_label}.xlsx"
    resp = make_response(payload)
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp


@bp.route("/admin/statistics/drug-outbound", methods=["GET"])
@role_required(['admin', 'nurse', 'finance'])
def get_drug_outbound_records():
    def parse_dt(value, is_end=False):
        return parse_local_datetime(value, is_end=is_end)

    start_time_str = request.args.get("start_time") or request.args.get("start_date")
    end_time_str = request.args.get("end_time") or request.args.get("end_date")
    if not start_time_str and not end_time_str:
        today = local_now().strftime("%Y-%m-%d")
        start_time_str = f"{today} 00:00:00"
        end_time_str = f"{today} 23:59:59"

    doctor_id = request.args.get("doctor_id", type=int)
    nurse_id = request.args.get("nurse_id", type=int)
    keyword = (request.args.get("keyword") or "").strip()
    page = request.args.get("page", 1, type=int)
    size = request.args.get("size", 50, type=int)
    if page <= 0:
        page = 1
    if size <= 0:
        size = 50
    if size > 200:
        size = 200

    try:
        start = parse_dt(start_time_str, is_end=False) if start_time_str else None
        end = parse_dt(end_time_str, is_end=True) if end_time_str else None
        if start is None and end is None:
            raise ValueError("Invalid date format")
        if start is None:
            start = end - timedelta(days=1)
        if end is None:
            end = start + timedelta(days=1)
    except Exception:
        return jsonify({"msg": "Invalid date format"}), 400

    Doctor = aliased(User)
    Nurse = aliased(User)

    amount_expr = func.coalesce(PrescriptionItem.new_amount, PrescriptionItem.amount)
    q = (
        db.session.query(
            Payment.payment_date,
            Payment.payment_method,
            Visit.id.label("visit_id"),
            Patient.name.label("patient_name"),
            Patient.student_id.label("student_id"),
            Doctor.real_name.label("doctor_name"),
            Nurse.real_name.label("nurse_name"),
            Drug.id.label("drug_id"),
            Drug.name.label("drug_name"),
            Drug.specification.label("specification"),
            Drug.unit.label("unit"),
            PrescriptionItem.is_scattered.label("is_scattered"),
            PrescriptionItem.quantity.label("quantity"),
            PrescriptionItem.price_at_visit.label("price_at_visit"),
            amount_expr.label("amount"),
        )
        .join(Visit, Payment.visit_id == Visit.id)
        .join(Patient, Visit.patient_id == Patient.id)
        .join(Doctor, Visit.doctor_id == Doctor.id)
        .join(Nurse, Payment.nurse_id == Nurse.id)
        .join(PrescriptionItem, PrescriptionItem.visit_id == Visit.id)
        .join(Drug, PrescriptionItem.drug_id == Drug.id)
        .filter(Payment.payment_date >= start, Payment.payment_date < end)
        .filter(Drug.type.in_([1, 3]))
    )

    if doctor_id:
        q = q.filter(Visit.doctor_id == doctor_id)
    if nurse_id:
        q = q.filter(Payment.nurse_id == nurse_id)
    if keyword:
        q = q.filter(func.lower(Drug.name).contains(keyword.lower()) | func.lower(Drug.specification).contains(keyword.lower()))

    total_records = q.count()
    agg = q.with_entities(
        func.coalesce(func.sum(amount_expr), 0.0),
        func.coalesce(func.sum(PrescriptionItem.quantity), 0),
    ).first()
    total_amount = float(agg[0] or 0.0) if agg else 0.0
    total_quantity = int(agg[1] or 0) if agg else 0

    rows = (
        q.order_by(Payment.payment_date.desc(), Visit.id.desc())
        .offset((page - 1) * size)
        .limit(size)
        .all()
    )

    details = []
    finance_view = _current_user_is_finance()
    for r in rows:
        details.append({
            "date": format_local_datetime(r.payment_date, "%Y-%m-%d %H:%M:%S") or "",
            "visit_id": r.visit_id,
            "patient_name": _mask_patient_name(r.patient_name, finance_view),
            "student_id": "" if finance_view else r.student_id,
            "doctor_name": r.doctor_name,
            "nurse_name": r.nurse_name,
            "payment_method": r.payment_method,
            "drug_id": r.drug_id,
            "drug_name": r.drug_name,
            "specification": r.specification,
            "unit": r.unit,
            "is_scattered": bool(r.is_scattered),
            "quantity": int(r.quantity or 0),
            "price_at_visit": float(r.price_at_visit or 0.0),
            "amount": float(r.amount or 0.0),
        })

    pages = int((total_records + size - 1) / size) if size else 1

    return jsonify({
        "data": {
            "summary": {
                "total_records": total_records,
                "total_quantity": total_quantity,
                "total_amount": total_amount,
            },
            "details": details,
            "meta": {
                "page": page,
                "per_page": size,
                "total": total_records,
                "pages": pages,
            },
            "range": {
                "start": utc_naive_to_local(start).strftime("%Y-%m-%d %H:%M:%S"),
                "end": utc_naive_to_local(end - timedelta(seconds=1)).strftime("%Y-%m-%d %H:%M:%S"),
            }
        }
    }), 200


@bp.route("/admin/statistics/drug-outbound/export", methods=["GET"])
@role_required(['admin', 'nurse', 'finance'])
def export_drug_outbound_records():
    from openpyxl import Workbook

    def parse_dt(value, is_end=False):
        return parse_local_datetime(value, is_end=is_end)

    def safe_text(val):
        s = "" if val is None else str(val)
        if s.startswith(("=", "+", "-", "@")):
            return "'" + s
        return s

    start_time_str = request.args.get("start_time") or request.args.get("start_date")
    end_time_str = request.args.get("end_time") or request.args.get("end_date")
    if not start_time_str and not end_time_str:
        today = local_now().strftime("%Y-%m-%d")
        start_time_str = f"{today} 00:00:00"
        end_time_str = f"{today} 23:59:59"

    doctor_id = request.args.get("doctor_id", type=int)
    nurse_id = request.args.get("nurse_id", type=int)
    keyword = (request.args.get("keyword") or "").strip()

    try:
        start = parse_dt(start_time_str, is_end=False) if start_time_str else None
        end = parse_dt(end_time_str, is_end=True) if end_time_str else None
        if start is None and end is None:
            raise ValueError("Invalid date format")
        if start is None:
            start = end - timedelta(days=1)
        if end is None:
            end = start + timedelta(days=1)
    except Exception:
        return jsonify({"msg": "Invalid date format"}), 400

    Doctor = aliased(User)
    Nurse = aliased(User)
    amount_expr = func.coalesce(PrescriptionItem.new_amount, PrescriptionItem.amount)
    q = (
        db.session.query(
            Payment.payment_date,
            Payment.payment_method,
            Visit.id.label("visit_id"),
            Patient.name.label("patient_name"),
            Patient.student_id.label("student_id"),
            Doctor.real_name.label("doctor_name"),
            Nurse.real_name.label("nurse_name"),
            Drug.name.label("drug_name"),
            Drug.specification.label("specification"),
            Drug.unit.label("unit"),
            PrescriptionItem.is_scattered.label("is_scattered"),
            PrescriptionItem.quantity.label("quantity"),
            PrescriptionItem.price_at_visit.label("price_at_visit"),
            amount_expr.label("amount"),
        )
        .join(Visit, Payment.visit_id == Visit.id)
        .join(Patient, Visit.patient_id == Patient.id)
        .join(Doctor, Visit.doctor_id == Doctor.id)
        .join(Nurse, Payment.nurse_id == Nurse.id)
        .join(PrescriptionItem, PrescriptionItem.visit_id == Visit.id)
        .join(Drug, PrescriptionItem.drug_id == Drug.id)
        .filter(Payment.payment_date >= start, Payment.payment_date < end)
        .filter(Drug.type.in_([1, 3]))
    )
    if doctor_id:
        q = q.filter(Visit.doctor_id == doctor_id)
    if nurse_id:
        q = q.filter(Payment.nurse_id == nurse_id)
    if keyword:
        q = q.filter(func.lower(Drug.name).contains(keyword.lower()) | func.lower(Drug.specification).contains(keyword.lower()))

    rows = q.order_by(Payment.payment_date.asc(), Visit.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "outbound"
    ws.append([
        "出库时间",
        "visit_id",
        "患者姓名",
        "学号",
        "接诊医生",
        "开药护士",
        "支付方式",
        "药品名称",
        "规格",
        "单位",
        "是否零散",
        "数量",
        "单价",
        "金额",
    ])

    total_amount = 0.0
    total_quantity = 0
    finance_view = _current_user_is_finance()
    for r in rows:
        qty = int(r.quantity or 0)
        amt = float(r.amount or 0.0)
        total_quantity += qty
        total_amount += amt
        ws.append([
            safe_text(format_local_datetime(r.payment_date, "%Y-%m-%d %H:%M:%S") or ""),
            r.visit_id,
            safe_text(_mask_patient_name(r.patient_name, finance_view)),
            safe_text("" if finance_view else r.student_id),
            safe_text(r.doctor_name),
            safe_text(r.nurse_name),
            safe_text(r.payment_method),
            safe_text(r.drug_name),
            safe_text(r.specification),
            safe_text(r.unit),
            "是" if bool(r.is_scattered) else "否",
            qty,
            round(float(r.price_at_visit or 0.0), 4),
            round(amt, 2),
        ])

    ws2 = wb.create_sheet("summary")
    local_start = utc_naive_to_local(start)
    local_end = utc_naive_to_local(end - timedelta(seconds=1))
    ws2.append(["开始时间", safe_text(local_start.strftime("%Y-%m-%d %H:%M:%S"))])
    ws2.append(["结束时间", safe_text(local_end.strftime("%Y-%m-%d %H:%M:%S"))])
    ws2.append(["doctor_id", safe_text(doctor_id if doctor_id else "")])
    ws2.append(["nurse_id", safe_text(nurse_id if nurse_id else "")])
    ws2.append(["keyword", safe_text(keyword)])
    ws2.append(["记录数", len(rows)])
    ws2.append(["合计数量", total_quantity])
    ws2.append(["合计金额", round(total_amount, 2)])

    stream = io.BytesIO()
    wb.save(stream)
    payload = stream.getvalue()

    filename = (
        f"drug_outbound_{local_start.strftime('%Y%m%d_%H%M')}_"
        f"{local_end.strftime('%Y%m%d_%H%M')}.xlsx"
    )
    resp = make_response(payload)
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp

@bp.route('/admin/users', methods=['GET'])
@role_required('admin')
def get_users():
    users = User.query.filter(User.role != 'admin').all()
    data = []
    for u in users:
        data.append({
            "id": u.id,
            "username": u.username,
            "real_name": u.real_name,
            "role": u.role,
            "is_active": u.is_active is not False,
        })
    return jsonify({"data": data}), 200

@bp.route('/admin/users', methods=['POST'])
@role_required('admin')
def create_user():
    data = request.get_json() or {}
    required_fields = ['username', 'password', 'real_name', 'role']
    for field in required_fields:
        if not data.get(field):
            return jsonify({"msg": f"Missing required field: {field}"}), 400

    if User.query.filter_by(username=data['username']).first():
        return jsonify({"msg": "Username already exists"}), 400

    if data['role'] not in ['doctor', 'nurse', 'finance']:
        return jsonify({"msg": "角色无效，有效角色: doctor/nurse/finance"}), 400
    try:
        validate_password(data['password'])
    except ValueError:
        return jsonify({"msg": "密码长度不能少于12位", "field": "password"}), 400

    user = User(
        username=data['username'],
        real_name=data['real_name'],
        role=data['role']
    )
    user.set_password(data['password'])
    db.session.add(user)
    db.session.commit()

    return jsonify({"data": {"id": user.id}}), 201

@bp.route('/admin/users/<int:id>', methods=['PUT'])
@role_required('admin')
def update_user(id):
    user = db.get_or_404(User, id)
    if user.role == 'admin':
        return jsonify({"msg": "Cannot modify admin user"}), 403

    data = request.get_json() or {}

    if 'username' in data and data['username'] != user.username:
        if User.query.filter_by(username=data['username']).first():
            return jsonify({"msg": "Username already exists"}), 400
        user.username = data['username']

    if 'real_name' in data:
        user.real_name = data['real_name']

    if 'role' in data and data['role'] in ['doctor', 'nurse', 'finance']:
        user.role = data['role']

    if 'is_active' in data:
        if type(data['is_active']) is not bool:
            return jsonify({"msg": "is_active must be a boolean", "field": "is_active"}), 400
        requested_active = data['is_active']
        current_active = user.is_active is not False
        if requested_active != current_active or user.is_active is None:
            user.is_active = requested_active
            user.token_version = int(user.token_version or 0) + 1

    if data.get('password'):
        try:
            validate_password(data['password'])
        except ValueError:
            return jsonify({"msg": "密码长度不能少于12位", "field": "password"}), 400
        user.set_password(data['password'])

    db.session.commit()
    return jsonify({"msg": "User updated successfully"}), 200

@bp.route('/admin/users/<int:id>', methods=['DELETE'])
@role_required('admin')
def disable_user(id):
    user = db.get_or_404(User, id)
    if user.role == 'admin':
        return jsonify({"msg": "Cannot disable admin user"}), 403

    user.is_active = False
    user.token_version = int(user.token_version or 0) + 1
    db.session.commit()
    return jsonify({"msg": "User disabled successfully"}), 200

@bp.route('/admin/operation-logs', methods=['GET'])
@role_required(['admin'])
def get_operation_logs():
    """获取运营日志列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('size', 20, type=int)
    action_type = request.args.get('action_type', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    query = OperationLog.query.order_by(OperationLog.timestamp.desc())
    
    if action_type:
        query = query.filter(OperationLog.action_type == action_type)
    
    if start_date:
        try:
            start_dt = parse_local_datetime(start_date)
            query = query.filter(OperationLog.timestamp >= start_dt)
        except ValueError:
            return jsonify({"msg": "Invalid start_date"}), 400
    
    if end_date:
        try:
            end_dt = parse_local_datetime(end_date, is_end=True)
            query = query.filter(OperationLog.timestamp < end_dt)
        except ValueError:
            return jsonify({"msg": "Invalid end_date"}), 400
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    data = []
    for log in pagination.items:
        user = log.user
        data.append({
            "id": log.id,
            "user_name": user.real_name if user else "未知",
            "user_role": user.role if user else "",
            "action_type": log.action_type,
            "target_type": log.target_type,
            "target_id": log.target_id,
            "summary": log.summary,
            "details": log.details,
            "timestamp": format_local_datetime(log.timestamp, "%Y-%m-%d %H:%M:%S") or ""
        })
    
    return jsonify({
        "data": data,
        "meta": {
            "page": page,
            "per_page": per_page,
            "total": pagination.total
        }
    }), 200

@bp.route('/admin/backup', methods=['GET', 'POST'])
@role_required('admin')
def backup_database():
    try:
        uri = current_app.config.get('SQLALCHEMY_DATABASE_URI', '')
        if uri.startswith('mysql'):
            return _mysql_backup_response(uri)

        backup_path, error_response = _create_sqlite_backup(uri)
        if error_response:
            return error_response

        if request.method == 'POST':
            _prune_sqlite_backups(backup_path)
            return jsonify({
                "msg": "Backup successful",
                "filename": os.path.basename(backup_path),
            }), 200

        try:
            response = send_file(
                backup_path,
                as_attachment=True,
                download_name=os.path.basename(backup_path),
                mimetype='application/vnd.sqlite3',
            )
        except Exception:
            os.remove(backup_path)
            raise
        response.direct_passthrough = False
        response.call_on_close(lambda: os.path.exists(backup_path) and os.remove(backup_path))
        return response
    except Exception as e:
        current_app.logger.exception("Database backup failed")
        return jsonify({"msg": "备份失败，请检查服务日志"}), 500
